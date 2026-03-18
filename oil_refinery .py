"""
oil_refinery - ERP
Single-file version — Python + Tkinter + SQLite
Run: python oil_refinery.py
Default Login: admin / admin123
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import sqlite3
import hashlib
import os

# ═══════════════════════════════════════════════════════════════════════════════
#  STYLES / THEME
# ═══════════════════════════════════════════════════════════════════════════════
BG_MAIN       = "#F0F4F8"
BG_SIDEBAR    = "#1A2332"
BG_TOPBAR     = "#FFFFFF"
BG_CARD       = "#FFFFFF"
ACCENT        = "#2ECC71"
ACCENT_DARK   = "#27AE60"
DANGER        = "#E74C3C"
DANGER_DARK   = "#C0392B"
WARNING       = "#F39C12"
WARNING_DARK  = "#D68910"
INFO          = "#2980B9"
INFO_DARK     = "#1F618D"
TEXT_MAIN     = "#2C3E50"
TEXT_MUTED    = "#7F8C8D"
TEXT_LIGHT    = "#FFFFFF"
TEXT_SIDEBAR  = "#BDC3C7"
SIDEBAR_ACTIVE_BG   = "#2ECC71"
SIDEBAR_ACTIVE_TEXT = "#FFFFFF"
SIDEBAR_HOVER_BG    = "#253446"
BORDER        = "#DDE3EA"
ENTRY_BG      = "#FFFFFF"
ENTRY_FG      = "#2C3E50"
TABLE_HEAD_BG = "#2C3E50"
TABLE_HEAD_FG = "#FFFFFF"
TABLE_ROW_ODD  = "#FFFFFF"
TABLE_ROW_EVEN = "#F7F9FC"
TABLE_SEL_BG   = "#D5F5E3"
TABLE_SEL_FG   = "#2C3E50"
FONT_FAMILY   = "Segoe UI"
FONT_TITLE    = (FONT_FAMILY, 18, "bold")
FONT_SUBTITLE = (FONT_FAMILY, 13, "bold")
FONT_NORMAL   = (FONT_FAMILY, 10)
FONT_SMALL    = (FONT_FAMILY, 9)
FONT_BOLD     = (FONT_FAMILY, 10, "bold")
FONT_SIDEBAR  = (FONT_FAMILY, 10)
SIDEBAR_W     = 210
TOPBAR_H      = 55
PAD           = 12
PAD_SMALL     = 6

try:
    from tkcalendar import DateEntry
    CAL_OK = True
except ImportError:
    CAL_OK = False

try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as rl_canvas
    PDF_OK = True
except ImportError:
    PDF_OK = False


# ═══════════════════════════════════════════════════════════════════════════════
#  THEME SYSTEM (Dark / Light)
# ═══════════════════════════════════════════════════════════════════════════════
THEMES = {
    "light": {
        "BG_MAIN": "#F0F4F8", "BG_SIDEBAR": "#1A2332", "BG_TOPBAR": "#FFFFFF",
        "BG_CARD": "#FFFFFF", "TEXT_MAIN": "#2C3E50", "TEXT_MUTED": "#7F8C8D",
        "TEXT_LIGHT": "#FFFFFF", "TEXT_SIDEBAR": "#BDC3C7",
        "SIDEBAR_ACTIVE_BG": "#2ECC71", "SIDEBAR_ACTIVE_TEXT": "#FFFFFF",
        "SIDEBAR_HOVER_BG": "#253446", "BORDER": "#DDE3EA",
        "ENTRY_BG": "#FFFFFF", "ENTRY_FG": "#2C3E50",
        "TABLE_HEAD_BG": "#2C3E50", "TABLE_HEAD_FG": "#FFFFFF",
        "TABLE_ROW_ODD": "#FFFFFF", "TABLE_ROW_EVEN": "#F7F9FC",
        "TABLE_SEL_BG": "#D5F5E3", "TABLE_SEL_FG": "#2C3E50",
    },
    "dark": {
        "BG_MAIN": "#1A1D2E", "BG_SIDEBAR": "#0F1117", "BG_TOPBAR": "#252836",
        "BG_CARD": "#252836", "TEXT_MAIN": "#E0E0E0", "TEXT_MUTED": "#8F9BB3",
        "TEXT_LIGHT": "#FFFFFF", "TEXT_SIDEBAR": "#8F9BB3",
        "SIDEBAR_ACTIVE_BG": "#2ECC71", "SIDEBAR_ACTIVE_TEXT": "#FFFFFF",
        "SIDEBAR_HOVER_BG": "#1E2130", "BORDER": "#353849",
        "ENTRY_BG": "#1E2130", "ENTRY_FG": "#E0E0E0",
        "TABLE_HEAD_BG": "#0F1117", "TABLE_HEAD_FG": "#E0E0E0",
        "TABLE_ROW_ODD": "#252836", "TABLE_ROW_EVEN": "#1E2130",
        "TABLE_SEL_BG": "#1A4731", "TABLE_SEL_FG": "#E0E0E0",
    }
}

_current_theme = "light"

def apply_theme(name):
    global _current_theme
    global BG_MAIN, BG_SIDEBAR, BG_TOPBAR, BG_CARD, TEXT_MAIN, TEXT_MUTED
    global TEXT_LIGHT, TEXT_SIDEBAR, SIDEBAR_ACTIVE_BG, SIDEBAR_ACTIVE_TEXT
    global SIDEBAR_HOVER_BG, BORDER, ENTRY_BG, ENTRY_FG
    global TABLE_HEAD_BG, TABLE_HEAD_FG, TABLE_ROW_ODD, TABLE_ROW_EVEN
    global TABLE_SEL_BG, TABLE_SEL_FG
    _current_theme = name
    t = THEMES[name]
    BG_MAIN = t["BG_MAIN"]; BG_SIDEBAR = t["BG_SIDEBAR"]
    BG_TOPBAR = t["BG_TOPBAR"]; BG_CARD = t["BG_CARD"]
    TEXT_MAIN = t["TEXT_MAIN"]; TEXT_MUTED = t["TEXT_MUTED"]
    TEXT_LIGHT = t["TEXT_LIGHT"]; TEXT_SIDEBAR = t["TEXT_SIDEBAR"]
    SIDEBAR_ACTIVE_BG = t["SIDEBAR_ACTIVE_BG"]
    SIDEBAR_ACTIVE_TEXT = t["SIDEBAR_ACTIVE_TEXT"]
    SIDEBAR_HOVER_BG = t["SIDEBAR_HOVER_BG"]; BORDER = t["BORDER"]
    ENTRY_BG = t["ENTRY_BG"]; ENTRY_FG = t["ENTRY_FG"]
    TABLE_HEAD_BG = t["TABLE_HEAD_BG"]; TABLE_HEAD_FG = t["TABLE_HEAD_FG"]
    TABLE_ROW_ODD = t["TABLE_ROW_ODD"]; TABLE_ROW_EVEN = t["TABLE_ROW_EVEN"]
    TABLE_SEL_BG = t["TABLE_SEL_BG"]; TABLE_SEL_FG = t["TABLE_SEL_FG"]


# ═══════════════════════════════════════════════════════════════════════════════
#  HARDWARE LOCK SYSTEM
# ═══════════════════════════════════════════════════════════════════════════════
def _get_machine_id() -> str:
    """Get unique machine hardware ID — works on Windows/Linux/Mac."""
    import platform, subprocess, uuid
    system = platform.system()
    try:
        if system == "Windows":
            # Windows: use CPU ID + Motherboard serial
            out = subprocess.check_output(
                "wmic csproduct get uuid", shell=True, stderr=subprocess.DEVNULL
            ).decode().strip().split("\n")[-1].strip()
            if out and len(out) > 5:
                return hashlib.md5(out.encode()).hexdigest()
        elif system == "Linux":
            with open("/etc/machine-id") as f:
                return hashlib.md5(f.read().strip().encode()).hexdigest()
        elif system == "Darwin":  # Mac
            out = subprocess.check_output(
                "ioreg -rd1 -c IOPlatformExpertDevice | grep IOPlatformUUID",
                shell=True, stderr=subprocess.DEVNULL
            ).decode()
            uid = out.split('"')[-2]
            return hashlib.md5(uid.encode()).hexdigest()
    except Exception:
        pass
    # Fallback: use MAC address
    mac = ':'.join(['{:02x}'.format((uuid.getnode() >> ele) & 0xff)
                    for ele in range(0, 48, 8)][::-1])
    return hashlib.md5(mac.encode()).hexdigest()


def _get_lock_file() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), ".hwlock")


def check_hardware_lock() -> bool:
    """
    Returns True if app is allowed to run on this machine.
    First run: registers this machine automatically.
    Subsequent runs: checks if machine matches.
    """
    lock_file = _get_lock_file()
    current_id = _get_machine_id()

    if not os.path.exists(lock_file):
        # First run — register this machine
        with open(lock_file, "w") as f:
            f.write(current_id)
        return True

    with open(lock_file, "r") as f:
        stored_id = f.read().strip()

    return current_id == stored_id


def show_lock_error():
    """Show unauthorized device error and exit."""
    root = tk.Tk(); root.withdraw()
    messagebox.showerror(
        "Unauthorized Device ❌",
        "Ye application is computer ke liye authorized nahi hai!\n\n"
        "This software is licensed for a specific device only.\n"
        "Please contact the software provider.\n\n"
        "Error Code: HW-LOCK-001"
    )
    root.destroy()
    import sys; sys.exit(1)


DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "oil_refinery.db")


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def initialize_database():
    conn = get_connection()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'admin',
        created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')))""")
    c.execute("""CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        base_unit TEXT NOT NULL,
        trade_unit TEXT,
        conversion_factor REAL NOT NULL DEFAULT 1.0,
        stock REAL NOT NULL DEFAULT 0.0,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')))""")
    c.execute("""CREATE TABLE IF NOT EXISTS parties (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        address TEXT,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')))""")
    c.execute("""CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        type TEXT NOT NULL CHECK(type IN ('SALE','PURCHASE','PAYMENT IN','PAYMENT OUT')),
        party_id INTEGER NOT NULL REFERENCES parties(id),
        product_id INTEGER NOT NULL REFERENCES products(id),
        quantity_base REAL NOT NULL,
        entered_quantity REAL NOT NULL,
        entered_unit TEXT NOT NULL,
        price_per_unit REAL NOT NULL,
        gst REAL NOT NULL DEFAULT 0.0,
        total_amount REAL NOT NULL,
        payment_cash REAL NOT NULL DEFAULT 0.0,
        payment_online REAL NOT NULL DEFAULT 0.0,
        balance_amount REAL NOT NULL,
        remarks TEXT,
        created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')))""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_txn_date    ON transactions(date)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_txn_party   ON transactions(party_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_txn_product ON transactions(product_id)")
    c.execute("""CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        position TEXT,
        salary REAL NOT NULL DEFAULT 0.0)""")
    c.execute("""CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        category TEXT NOT NULL,
        employee_id INTEGER REFERENCES employees(id),
        description TEXT,
        amount REAL NOT NULL,
        created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')))""")
    c.execute("CREATE INDEX IF NOT EXISTS idx_exp_date ON expenses(date)")
    c.execute("""CREATE TABLE IF NOT EXISTS production_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')))""")
    c.execute("""CREATE TABLE IF NOT EXISTS production_inputs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id INTEGER NOT NULL REFERENCES production_batches(id),
        product_id INTEGER NOT NULL REFERENCES products(id),
        quantity_base REAL NOT NULL,
        entered_quantity REAL NOT NULL,
        entered_unit TEXT NOT NULL)""")
    c.execute("""CREATE TABLE IF NOT EXISTS production_outputs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id INTEGER NOT NULL REFERENCES production_batches(id),
        product_id INTEGER NOT NULL REFERENCES products(id),
        quantity_base REAL NOT NULL,
        entered_quantity REAL NOT NULL,
        entered_unit TEXT NOT NULL)""")
    # Settings table for email backup config
    c.execute("""CREATE TABLE IF NOT EXISTS settings (
        key   TEXT PRIMARY KEY,
        value TEXT NOT NULL DEFAULT '')""")
    # Default settings
    for key, val in [
        ("backup_email",      ""),
        ("sender_email",      ""),
        ("sender_password",   ""),
        ("auto_email_backup", "0"),
    ]:
        c.execute("INSERT OR IGNORE INTO settings (key,value) VALUES (?,?)", (key,val))
    conn.commit()
    conn.close()


def get_setting(key, default=""):
    try:
        conn = get_connection()
        row  = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        conn.close()
        return row["value"] if row else default
    except: return default

def set_setting(key, value):
    try:
        conn = get_connection()
        conn.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (key,value))
        conn.commit(); conn.close()
    except: pass

def send_email_backup(backup_path):
    """
    Backup file email pe bhejo.
    Returns (True, "") on success or (False, error_msg) on failure.
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base      import MIMEBase
    from email.mime.text      import MIMEText
    from email                import encoders

    sender    = get_setting("sender_email")
    password  = get_setting("sender_password")
    recipient = get_setting("backup_email")

    if not sender or not password or not recipient:
        return False, "Email settings puri nahi hain!\nSettings → Email Backup mein jaake fill karein."

    try:
        msg = MIMEMultipart()
        msg["From"]    = sender
        msg["To"]      = recipient
        msg["Subject"] = f"Saark Industries — DB Backup {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        body = (f"Saark Industries ERP — Automatic Backup\n\n"
                f"Date: {datetime.now().strftime('%d %B %Y, %I:%M %p')}\n"
                f"File: {os.path.basename(backup_path)}\n\n"
                f"Ye email automatically generate hui hai.\n"
                f"Backup file attachment mein hai.")
        msg.attach(MIMEText(body, "plain"))

        # Attach backup file
        with open(backup_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f"attachment; filename={os.path.basename(backup_path)}")
        msg.attach(part)

        # Send via Gmail SMTP
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as server:
            server.login(sender, password)
            server.sendmail(sender, recipient, msg.as_string())

        return True, ""
    except smtplib.SMTPAuthenticationError:
        return False, ("Gmail login fail!\n"
                       "App Password sahi se daala?\n"
                       "Guide: myaccount.google.com → Security → App Passwords")
    except smtplib.SMTPException as e:
        return False, f"Email send nahi hua: {e}"
    except Exception as e:
        return False, f"Error: {e}"


def today_str():
    return datetime.now().strftime("%Y-%m-%d")


# ═══════════════════════════════════════════════════════════════════════════════
#  REUSABLE WIDGETS
# ═══════════════════════════════════════════════════════════════════════════════
def make_date_entry(parent, var):
    if CAL_OK:
        return DateEntry(parent, textvariable=var, date_pattern="yyyy-mm-dd",
                         font=FONT_NORMAL, background=ACCENT, foreground="white", width=14)
    return tk.Entry(parent, textvariable=var, font=FONT_NORMAL,
                    bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1, width=16)


class StyledButton(tk.Button):
    COLORS = {
        "primary": (ACCENT,   ACCENT_DARK,  TEXT_LIGHT),
        "danger":  (DANGER,   DANGER_DARK,  TEXT_LIGHT),
        "warning": (WARNING,  WARNING_DARK, TEXT_LIGHT),
        "info":    (INFO,     INFO_DARK,    TEXT_LIGHT),
        "neutral": ("#95A5A6","#7F8C8D",    TEXT_LIGHT),
    }
    def __init__(self, parent, text, command=None, kind="primary", **kw):
        bg, hover, fg = self.COLORS.get(kind, self.COLORS["primary"])
        super().__init__(parent, text=text, command=command,
                         bg=bg, fg=fg, font=FONT_BOLD,
                         relief="flat", bd=0, cursor="hand2",
                         padx=14, pady=6,
                         activebackground=hover, activeforeground=fg, **kw)
        self._bg = bg; self._hv = hover
        self.bind("<Enter>", lambda e: self.config(bg=self._hv))
        self.bind("<Leave>", lambda e: self.config(bg=self._bg))


class StyledTable(tk.Frame):
    def __init__(self, parent, columns, **kw):
        super().__init__(parent, bg=BG_CARD, **kw)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("ST.Treeview", background=TABLE_ROW_ODD, foreground=TEXT_MAIN,
                         rowheight=28, fieldbackground=TABLE_ROW_ODD, font=FONT_NORMAL, borderwidth=0)
        style.configure("ST.Treeview.Heading", background=TABLE_HEAD_BG,
                         foreground=TABLE_HEAD_FG, font=FONT_BOLD, relief="flat")
        style.map("ST.Treeview",
                  background=[("selected", TABLE_SEL_BG)],
                  foreground=[("selected", TABLE_SEL_FG)])
        col_ids = [c[0] for c in columns]
        self.tree = ttk.Treeview(self, columns=col_ids, show="headings",
                                 style="ST.Treeview", selectmode="browse")
        for col_id, heading, width in columns:
            self.tree.heading(col_id, text=heading, anchor="w")
            self.tree.column(col_id, width=width, anchor="w", stretch=True)
        self.tree.tag_configure("even", background=TABLE_ROW_EVEN)
        self.tree.tag_configure("odd",  background=TABLE_ROW_ODD)
        vsb = ttk.Scrollbar(self, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1); self.columnconfigure(0, weight=1)
        self._all_rows = []  # Store all rows for search filtering

    def load(self, rows):
        self._all_rows = list(rows)
        self._render(rows)

    def _render(self, rows):
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows):
            self.tree.insert("", "end", values=row, tags=("even" if i%2==0 else "odd",))

    def search(self, keyword):
        """Filter rows by keyword — searches all columns."""
        kw = keyword.strip().lower()
        if not kw:
            self._render(self._all_rows)
            return
        filtered = [r for r in self._all_rows
                    if any(kw in str(v).lower() for v in r)]
        self._render(filtered)

    def get_selected(self):
        sel = self.tree.selection()
        return self.tree.item(sel[0])["values"] if sel else None


class SearchBar(tk.Frame):
    """
    Reusable live search bar — type karo, table apne aap filter ho jaye.
    Usage: SearchBar(parent, table=self.table, placeholder="Search...")
    """
    def __init__(self, parent, table: StyledTable, placeholder="🔍  Search karein...", **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        self._table = table
        self._var = tk.StringVar()
        self._var.trace_add("write", self._on_change)

        # Search icon + entry in a styled frame
        container = tk.Frame(self, bg=ENTRY_BG, highlightthickness=1,
                             highlightbackground=ACCENT, padx=4, pady=2)
        container.pack(fill="x")
        tk.Label(container, text="🔍", font=FONT_NORMAL,
                 bg=ENTRY_BG, fg=ACCENT).pack(side="left", padx=(4,2))
        self._entry = tk.Entry(container, textvariable=self._var,
                               font=FONT_NORMAL, bg=ENTRY_BG, fg=TEXT_MAIN,
                               relief="flat", bd=0, width=40)
        self._entry.pack(side="left", fill="x", expand=True, ipady=4)
        self._entry.insert(0, placeholder)
        self._entry.config(fg=TEXT_MUTED)
        self._entry.bind("<FocusIn>",  self._on_focus_in)
        self._entry.bind("<FocusOut>", self._on_focus_out)
        self._placeholder = placeholder

        # Clear button
        clear_btn = tk.Button(container, text="✕", font=FONT_SMALL,
                              bg=ENTRY_BG, fg=TEXT_MUTED, relief="flat",
                              bd=0, cursor="hand2", command=self._clear)
        clear_btn.pack(side="right", padx=4)

        # Result count label
        self._count_lbl = tk.Label(self, text="", font=FONT_SMALL,
                                   bg=BG_MAIN, fg=TEXT_MUTED)
        self._count_lbl.pack(anchor="e", padx=4)

    def _on_focus_in(self, _):
        if self._entry.get() == self._placeholder:
            self._entry.delete(0, "end")
            self._entry.config(fg=TEXT_MAIN)

    def _on_focus_out(self, _):
        if not self._entry.get():
            self._entry.insert(0, self._placeholder)
            self._entry.config(fg=TEXT_MUTED)

    def _on_change(self, *_):
        kw = self._var.get()
        if kw == self._placeholder: kw = ""
        self._table.search(kw)
        # Update count
        total    = len(self._table._all_rows)
        children = len(self._table.tree.get_children())
        if kw and kw != self._placeholder:
            self._count_lbl.config(
                text=f"{children} results mili / {total} mein se",
                fg=ACCENT if children > 0 else DANGER)
        else:
            self._count_lbl.config(text=f"Total: {total} records", fg=TEXT_MUTED)

    def _clear(self):
        self._var.set("")
        self._entry.delete(0, "end")
        self._entry.insert(0, self._placeholder)
        self._entry.config(fg=TEXT_MUTED)
        self._table.search("")
        self._count_lbl.config(text="")

    def reset(self):
        """Call after load_data to reset search."""
        self._clear()




class SectionHeader(tk.Frame):
    def __init__(self, parent, title, **kw):
        super().__init__(parent, bg=BG_MAIN, **kw)
        tk.Frame(self, height=3, bg=ACCENT).pack(fill="x")
        tk.Label(self, text=title, font=FONT_TITLE, bg=BG_MAIN,
                 fg=TEXT_MAIN, padx=PAD, pady=8).pack(anchor="w")


class ModalDialog(tk.Toplevel):
    def __init__(self, parent, title, width=500, height=420):
        super().__init__(parent)
        self.title(title)
        self.resizable(True, True)
        self.minsize(width, min(height, 400))
        self.configure(bg=BG_CARD)
        self.grab_set()
        # Screen ke hisaab se size adjust karo
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = min(width,  sw - 40)
        h = min(height, sh - 80)
        px = parent.winfo_rootx() + parent.winfo_width()  // 2 - w // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - h // 2
        # Screen ke bahar na jaye
        px = max(10, min(px, sw - w - 10))
        py = max(10, min(py, sh - h - 10))
        self.geometry(f"{w}x{h}+{px}+{py}")
        tk.Frame(self, height=4, bg=ACCENT).pack(fill="x")
        tk.Label(self, text=title, font=FONT_SUBTITLE, bg=BG_CARD,
                 fg=TEXT_MAIN, padx=PAD, pady=PAD).pack(anchor="w")
        tk.Frame(self, height=1, bg=BORDER).pack(fill="x")
        self.body = tk.Frame(self, bg=BG_CARD, padx=PAD, pady=PAD)
        self.body.pack(fill="both", expand=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  FIRST TIME SETUP WINDOW
# ═══════════════════════════════════════════════════════════════════════════════
class FirstSetupWindow(tk.Tk):
    """
    Pehli baar app khulne pe yahan aao —
    koi default user nahi hoga, khud account banao.
    """
    def __init__(self):
        super().__init__()
        self.title("Saark Industries - First Time Setup")
        self.geometry("460x650")
        self.resizable(True, True)
        self.configure(bg=BG_MAIN)
        self.setup_done = False
        self._build()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"460x650+{(sw-460)//2}+{(sh-650)//2}")

    def _build(self):
        # Banner
        banner = tk.Frame(self, bg=BG_SIDEBAR, height=160)
        banner.pack(fill="x"); banner.pack_propagate(False)
        tk.Label(banner, text="🏭", font=(FONT_FAMILY, 42),
                 bg=BG_SIDEBAR, fg=ACCENT).pack(pady=(20,4))
        tk.Label(banner, text="Saark Industries",
                 font=(FONT_FAMILY, 16, "bold"),
                 bg=BG_SIDEBAR, fg=TEXT_LIGHT).pack()
        tk.Label(banner, text="Pehli Baar Setup",
                 font=FONT_NORMAL, bg=BG_SIDEBAR, fg=TEXT_SIDEBAR).pack()

        card = tk.Frame(self, bg=BG_CARD, padx=32, pady=20)
        card.pack(fill="both", expand=True, padx=24, pady=20)

        # Welcome message
        tk.Label(card, text="👋 Welcome!",
                 font=FONT_SUBTITLE, bg=BG_CARD, fg=TEXT_MAIN).pack(anchor="w")
        tk.Label(card,
                 text="Apna Admin account banao.\n"
                      "Ye account sirf aapka hoga — koi default password nahi!",
                 font=FONT_SMALL, bg=BG_CARD, fg=TEXT_MUTED,
                 justify="left").pack(anchor="w", pady=(4, 16))

        tk.Frame(card, height=1, bg=BORDER).pack(fill="x", pady=(0,16))

        # Username
        tk.Label(card, text="Username *", font=FONT_SMALL,
                 bg=BG_CARD, fg=TEXT_MUTED).pack(anchor="w")
        self.user_var = tk.StringVar()
        tk.Entry(card, textvariable=self.user_var, font=FONT_NORMAL,
                 bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1
                 ).pack(fill="x", ipady=6, pady=(2, 12))

        # Password
        tk.Label(card, text="Password *", font=FONT_SMALL,
                 bg=BG_CARD, fg=TEXT_MUTED).pack(anchor="w")
        self.pass_var = tk.StringVar()
        tk.Entry(card, textvariable=self.pass_var, show="•",
                 font=FONT_NORMAL, bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="solid", bd=1
                 ).pack(fill="x", ipady=6, pady=(2, 12))

        # Confirm Password
        tk.Label(card, text="Confirm Password *", font=FONT_SMALL,
                 bg=BG_CARD, fg=TEXT_MUTED).pack(anchor="w")
        self.conf_var = tk.StringVar()
        conf_e = tk.Entry(card, textvariable=self.conf_var, show="•",
                          font=FONT_NORMAL, bg=ENTRY_BG, fg=ENTRY_FG,
                          relief="solid", bd=1)
        conf_e.pack(fill="x", ipady=6, pady=(2, 16))
        conf_e.bind("<Return>", lambda e: self._create())

        # Info box
        info = tk.Frame(card, bg="#EBF5FB", padx=10, pady=8)
        info.pack(fill="x", pady=(0, 16))
        tk.Label(info,
                 text="💡 Ye username/password yaad rakhein!\n"
                      "Baad mein aur users bhi add kar sakte hain.",
                 font=FONT_SMALL, bg="#EBF5FB", fg=INFO,
                 justify="left").pack(anchor="w")

        # Create button
        btn = tk.Button(card, text="✅ Admin Account Banao",
                        command=self._create,
                        bg=ACCENT, fg=TEXT_LIGHT, font=FONT_BOLD,
                        relief="flat", bd=0, cursor="hand2", pady=10)
        btn.pack(fill="x")
        btn.bind("<Enter>", lambda e: btn.config(bg=ACCENT_DARK))
        btn.bind("<Leave>", lambda e: btn.config(bg=ACCENT))

    def _create(self):
        uname = self.user_var.get().strip()
        pwd   = self.pass_var.get().strip()
        conf  = self.conf_var.get().strip()

        if not uname:
            messagebox.showerror("Error", "Username zaroori hai!", parent=self); return
        if len(uname) < 3:
            messagebox.showerror("Error", "Username kam se kam 3 characters ka hona chahiye!", parent=self); return
        if not pwd:
            messagebox.showerror("Error", "Password zaroori hai!", parent=self); return
        if len(pwd) < 6:
            messagebox.showerror("Error", "Password kam se kam 6 characters ka hona chahiye!", parent=self); return
        if pwd != conf:
            messagebox.showerror("Error", "Dono passwords match nahi karte!", parent=self); return

        try:
            conn = get_connection()
            conn.execute(
                "INSERT INTO users (username, password, role) VALUES (?,?,?)",
                (uname, hash_password(pwd), "admin")
            )
            conn.commit(); conn.close()
            messagebox.showinfo("Success! ✅",
                                f"Admin account ban gaya!\n\n"
                                f"Username: {uname}\n\n"
                                f"Ab login karein.",
                                parent=self)
            self.setup_done = True
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ═══════════════════════════════════════════════════════════════════════════════
class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Saark Industries - Login")
        self.geometry("420x520")
        self.resizable(False, False)
        self.configure(bg=BG_MAIN)
        self.logged_in_user = None
        self._build()
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"420x520+{(sw-420)//2}+{(sh-520)//2}")

    def _build(self):
        banner = tk.Frame(self, bg=BG_SIDEBAR, height=180)
        banner.pack(fill="x"); banner.pack_propagate(False)
        tk.Label(banner, text="⚙", font=(FONT_FAMILY, 48), bg=BG_SIDEBAR, fg=ACCENT).pack(pady=(30,4))
        tk.Label(banner, text="Saark Industries", font=(FONT_FAMILY,16,"bold"), bg=BG_SIDEBAR, fg=TEXT_LIGHT).pack()
        tk.Label(banner, text="Management System", font=FONT_NORMAL, bg=BG_SIDEBAR, fg=TEXT_SIDEBAR).pack()

        card = tk.Frame(self, bg=BG_CARD, padx=32, pady=24)
        card.pack(fill="both", expand=True, padx=24, pady=24)
        tk.Label(card, text="Sign In", font=FONT_SUBTITLE, bg=BG_CARD, fg=TEXT_MAIN).pack(anchor="w", pady=(0,16))

        tk.Label(card, text="Username", font=FONT_SMALL, bg=BG_CARD, fg=TEXT_MUTED).pack(anchor="w")
        self.user_var = tk.StringVar()
        tk.Entry(card, textvariable=self.user_var, font=FONT_NORMAL, bg=ENTRY_BG, fg=ENTRY_FG,
                 relief="solid", bd=1).pack(fill="x", ipady=6, pady=(2,12))

        tk.Label(card, text="Password", font=FONT_SMALL, bg=BG_CARD, fg=TEXT_MUTED).pack(anchor="w")
        self.pass_var = tk.StringVar()
        pw = tk.Entry(card, textvariable=self.pass_var, show="•", font=FONT_NORMAL,
                      bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1)
        pw.pack(fill="x", ipady=6, pady=(2,20))
        pw.bind("<Return>", lambda e: self._login())

        btn = tk.Button(card, text="LOGIN", command=self._login, bg=ACCENT, fg=TEXT_LIGHT,
                        font=FONT_BOLD, relief="flat", bd=0, cursor="hand2", pady=10)
        btn.pack(fill="x")
        btn.bind("<Enter>", lambda e: btn.config(bg=ACCENT_DARK))
        btn.bind("<Leave>", lambda e: btn.config(bg=ACCENT))

    def _login(self):
        u = self.user_var.get().strip()
        p = self.pass_var.get().strip()
        if not u or not p:
            messagebox.showwarning("Login", "Enter username and password.", parent=self); return
        conn = get_connection()
        row = conn.execute("SELECT * FROM users WHERE username=? AND password=?",
                           (u, hash_password(p))).fetchone()
        conn.close()
        if row:
            self.logged_in_user = dict(row); self.destroy()
        else:
            messagebox.showerror("Failed", "Invalid username or password.", parent=self)
            self.pass_var.set("")


# ═══════════════════════════════════════════════════════════════════════════════
#  CHANGE PASSWORD DIALOG
# ═══════════════════════════════════════════════════════════════════════════════
class ChangePasswordDialog(ModalDialog):
    def __init__(self, parent, user):
        super().__init__(parent, "Change Password", 360, 320)
        self.user = user
        self._build()

    def _build(self):
        b = self.body; b.columnconfigure(1, weight=1)
        self.vars = {}
        for i, (lbl, key) in enumerate([("Current Password","old"),("New Password","new"),("Confirm Password","confirm")]):
            tk.Label(b, text=lbl, font=FONT_SMALL, bg=BG_CARD, fg=TEXT_MUTED
                     ).grid(row=i, column=0, sticky="w", pady=8, padx=(0,12))
            v = tk.StringVar(); self.vars[key] = v
            tk.Entry(b, textvariable=v, show="•", font=FONT_NORMAL,
                     bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1
                     ).grid(row=i, column=1, sticky="ew", ipady=4)
        r = tk.Frame(b, bg=BG_CARD); r.grid(row=3, column=0, columnspan=2, pady=16)
        StyledButton(r, "💾 Update", command=self._save).pack(side="left", padx=(0,8))
        StyledButton(r, "Cancel", command=self.destroy, kind="neutral").pack(side="left")

    def _save(self):
        old = self.vars["old"].get(); new = self.vars["new"].get(); confirm = self.vars["confirm"].get()
        if not old or not new: messagebox.showerror("Error","All fields required.", parent=self); return
        if new != confirm:     messagebox.showerror("Error","Passwords don't match.", parent=self); return
        if len(new) < 6:       messagebox.showerror("Error","Min 6 characters.", parent=self); return
        conn = get_connection()
        row = conn.execute("SELECT id FROM users WHERE id=? AND password=?",
                           (self.user["id"], hash_password(old))).fetchone()
        if not row:
            messagebox.showerror("Error","Current password wrong.", parent=self)
            conn.close(); return
        conn.execute("UPDATE users SET password=? WHERE id=?", (hash_password(new), self.user["id"]))
        conn.commit(); conn.close()
        messagebox.showinfo("Success","Password changed!", parent=self)
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
class DashboardFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._build()
        self.refresh()

    def _build(self):
        top = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=PAD)
        top.pack(fill="x")
        tk.Label(top, text="Dashboard", font=FONT_TITLE, bg=BG_MAIN, fg=TEXT_MAIN).pack(side="left")
        tk.Label(top, text=datetime.now().strftime("%A, %d %B %Y"),
                 font=FONT_NORMAL, bg=BG_MAIN, fg=TEXT_MUTED).pack(side="right", padx=8)
        StyledButton(top, "🔄 Refresh", command=self.refresh, kind="info").pack(side="right")

        cards = tk.Frame(self, bg=BG_MAIN, padx=PAD)
        cards.pack(fill="x", pady=(0,PAD))
        self._kpi = {}
        for col, (title, color, key) in enumerate([
            ("Total Receivable", DANGER,  "recv"),
            ("Total Payable",    INFO,    "pay"),
            ("Total Products",   ACCENT,  "prods"),
            ("Today Txns",       WARNING, "today"),
        ]):
            f = tk.Frame(cards, bg=color, padx=16, pady=14)
            f.grid(row=0, column=col, padx=6, sticky="ew")
            cards.columnconfigure(col, weight=1)
            tk.Label(f, text=title, font=FONT_SMALL, bg=color, fg="white").pack()
            lbl = tk.Label(f, text="—", font=(FONT_FAMILY,20,"bold"), bg=color, fg="white")
            lbl.pack()
            self._kpi[key] = lbl

        split = tk.Frame(self, bg=BG_MAIN, padx=PAD)
        split.pack(fill="both", expand=True)
        split.columnconfigure(0, weight=1); split.columnconfigure(1, weight=1)

        left = tk.LabelFrame(split, text=" Party-wise Outstanding ", font=FONT_BOLD,
                             bg=BG_CARD, fg=TEXT_MAIN, padx=6, pady=6)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        self.party_table = StyledTable(left, [("party","Party Name",200),("recv","Receivable",110),
                                               ("pay","Payable",110),("net","Net",110)])
        self.party_table.pack(fill="both", expand=True)

        right = tk.LabelFrame(split, text=" Product-wise Stock ", font=FONT_BOLD,
                              bg=BG_CARD, fg=TEXT_MAIN, padx=6, pady=6)
        right.grid(row=0, column=1, sticky="nsew", padx=(6,0))
        st = tk.Frame(right, bg=BG_CARD); st.pack(fill="x", pady=(0,4))
        tk.Label(st, text="As of:", font=FONT_SMALL, bg=BG_CARD, fg=TEXT_MUTED).pack(side="left")
        self.stock_date_var = tk.StringVar(value=today_str())
        make_date_entry(st, self.stock_date_var).pack(side="left", padx=4)
        StyledButton(st, "Show", command=self._load_stock).pack(side="left", padx=4)
        self.stock_table = StyledTable(right, [("prod","Product",200),("stock","Stock",110),("unit","Unit",90)])
        self.stock_table.pack(fill="both", expand=True)

    def refresh(self):
        conn = get_connection()
        recv     = conn.execute("SELECT COALESCE(SUM(balance_amount),0) FROM transactions WHERE type IN ('SALE','PAYMENT OUT')").fetchone()[0]
        payable  = conn.execute("SELECT COALESCE(SUM(balance_amount),0) FROM transactions WHERE type IN ('PURCHASE','PAYMENT IN')").fetchone()[0]
        prods    = conn.execute("SELECT COUNT(*) FROM products WHERE is_active=1").fetchone()[0]
        today_c  = conn.execute("SELECT COUNT(*) FROM transactions WHERE date=?", (today_str(),)).fetchone()[0]
        parties  = conn.execute("""
            SELECT p.name,
                   COALESCE(SUM(CASE WHEN t.type IN ('SALE','PAYMENT OUT')    THEN t.balance_amount ELSE 0 END),0) as r,
                   COALESCE(SUM(CASE WHEN t.type IN ('PURCHASE','PAYMENT IN') THEN t.balance_amount ELSE 0 END),0) as py
            FROM parties p LEFT JOIN transactions t ON t.party_id=p.id
            WHERE p.is_active=1 GROUP BY p.id ORDER BY p.name""").fetchall()
        conn.close()
        self._kpi["recv"].config(text=f"₹{recv:,.0f}")
        self._kpi["pay"].config(text=f"₹{payable:,.0f}")
        self._kpi["prods"].config(text=str(prods))
        self._kpi["today"].config(text=str(today_c))
        self.party_table.load([(r["name"], f"₹{r['r']:,.0f}", f"₹{r['py']:,.0f}",
                                 f"₹{r['r']-r['py']:,.0f}") for r in parties])
        self._load_stock()

    def _load_stock(self):
        as_of = self.stock_date_var.get()
        conn  = get_connection()
        prods = conn.execute("SELECT id,name,base_unit,stock FROM products WHERE is_active=1 ORDER BY name").fetchall()
        data  = []
        for p in prods:
            fs = conn.execute("SELECT COALESCE(SUM(quantity_base),0) FROM transactions WHERE product_id=? AND type='SALE' AND date>?",     (p["id"],as_of)).fetchone()[0]
            fp = conn.execute("SELECT COALESCE(SUM(quantity_base),0) FROM transactions WHERE product_id=? AND type='PURCHASE' AND date>?", (p["id"],as_of)).fetchone()[0]
            fo = conn.execute("SELECT COALESCE(SUM(po.quantity_base),0) FROM production_outputs po JOIN production_batches pb ON pb.id=po.batch_id WHERE po.product_id=? AND pb.date>?", (p["id"],as_of)).fetchone()[0]
            fi = conn.execute("SELECT COALESCE(SUM(pi.quantity_base),0) FROM production_inputs  pi JOIN production_batches pb ON pb.id=pi.batch_id WHERE pi.product_id=? AND pb.date>?", (p["id"],as_of)).fetchone()[0]
            stock_as_of = p["stock"] + fs - fp + fi - fo
            data.append((p["name"], f"{stock_as_of:.2f}", p["base_unit"]))
        conn.close()
        self.stock_table.load(data)


# ═══════════════════════════════════════════════════════════════════════════════
#  PRODUCTS
# ═══════════════════════════════════════════════════════════════════════════════
class ProductsFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._build(); self.load_data()

    def _build(self):
        SectionHeader(self, "Products Management").pack(fill="x")
        bar = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); bar.pack(fill="x")
        StyledButton(bar, "+ Add Product", command=self._add).pack(side="left")
        if PDF_OK: StyledButton(bar, "📄 PDF", command=self._pdf, kind="danger").pack(side="left", padx=8)
        self.table = StyledTable(self, [("id","ID",50),("name","Product Name",200),
            ("bu","Base Unit",100),("tu","Trade Unit",100),("cf","Conv.Factor",100),
            ("stock","Stock",120),("act","Active",60)])
        # Search bar
        self.search_bar = SearchBar(self, self.table, "🔍  Product name se search karein...")
        self.search_bar.pack(fill="x", padx=PAD, pady=(4,0))
        self.table.pack(fill="both", expand=True, padx=PAD, pady=PAD_SMALL)
        b = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); b.pack(fill="x")
        StyledButton(b, "✏ Edit",   command=self._edit,   kind="warning").pack(side="left", padx=(0,8))
        StyledButton(b, "🗑 Delete", command=self._delete, kind="danger").pack(side="left")

    def load_data(self):
        conn = get_connection()
        rows = conn.execute("SELECT id,name,base_unit,trade_unit,conversion_factor,stock,is_active FROM products ORDER BY name").fetchall()
        conn.close()
        self.table.load([(r["id"],r["name"],r["base_unit"],r["trade_unit"] or "–",
                          r["conversion_factor"],f"{r['stock']:.2f}","Yes" if r["is_active"] else "No") for r in rows])

    def _add(self):    ProductDialog(self, on_save=self.load_data)
    def _edit(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Edit","Select a product."); return
        conn = get_connection()
        row = conn.execute("SELECT * FROM products WHERE id=?", (sel[0],)).fetchone()
        conn.close()
        if row: ProductDialog(self, record=dict(row), on_save=self.load_data)
    def _delete(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Delete","Select a product."); return
        if not messagebox.askyesno("Delete", f"Delete '{sel[1]}'?", parent=self): return
        conn = get_connection()
        conn.execute("UPDATE products SET is_active=0 WHERE id=?", (sel[0],))
        conn.commit(); conn.close(); self.load_data()

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        rows = self.table._all_rows
        if not rows: messagebox.showinfo("PDF","Koi product nahi hai."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if not path: return
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        c.setFillColor("#1A2332"); c.rect(0, h-70, w, 70, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h-38, "Saark Industries — Products List")
        c.setFont("Helvetica", 9)
        c.drawString(40, h-56, f"Total Products: {len(rows)}   |   Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}")
        y = h - 90
        col_x = [40, 80, 230, 300, 360, 420, 490, 545]
        hdrs  = ["ID","Product Name","Base Unit","Trade Unit","Conv.Factor","Stock","Active"]
        c.setFillColor("#2C3E50"); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 8)
        for i, hdr in enumerate(hdrs):
            if i < len(col_x): c.drawString(col_x[i], y+2, hdr)
        y -= 18
        c.setFont("Helvetica", 8)
        for ri, row in enumerate(rows):
            if y < 60: c.showPage(); y = h-60; c.setFont("Helvetica", 8)
            bg = "#F7F9FC" if ri%2==0 else "white"
            c.setFillColor(bg); c.rect(38, y-2, w-76, 16, fill=1, stroke=0)
            c.setFillColor("#2C3E50")
            vals = [str(row[0]), str(row[1])[:20], str(row[2]), str(row[3]),
                    str(row[4]), str(row[5]), str(row[6])]
            for i, val in enumerate(vals):
                if i < len(col_x): c.drawString(col_x[i], y+1, val[:18])
            y -= 16
        c.save()
        messagebox.showinfo("PDF", f"PDF saved!\n{path}")


class ProductDialog(ModalDialog):
    def __init__(self, parent, record=None, on_save=None):
        super().__init__(parent, "Add Product" if not record else "Edit Product", 480, 360)
        self.record = record; self.on_save = on_save
        self._build()
        if record: self._populate()

    def _build(self):
        b = self.body; b.columnconfigure(1, weight=1)
        self.vars = {}
        for i, (lbl, key) in enumerate([("Product Name *","name"),("Base Unit *","bunit"),
                                          ("Trade Unit","tunit"),("Conv. Factor","cfact")]):
            tk.Label(b, text=lbl, font=FONT_NORMAL, bg=BG_CARD, fg=TEXT_MAIN
                     ).grid(row=i, column=0, sticky="w", pady=6, padx=(0,12))
            v = tk.StringVar(); self.vars[key] = v
            tk.Entry(b, textvariable=v, font=FONT_NORMAL, bg=ENTRY_BG, fg=ENTRY_FG,
                     relief="solid", bd=1).grid(row=i, column=1, sticky="ew", ipady=4)
        self.vars["cfact"].set("1.0")
        r = tk.Frame(b, bg=BG_CARD); r.grid(row=4, column=0, columnspan=2, pady=16)
        StyledButton(r, "💾 Save", command=self._save).pack(side="left", padx=(0,8))
        StyledButton(r, "Cancel", command=self.destroy, kind="neutral").pack(side="left")

    def _populate(self):
        self.vars["name"].set(self.record["name"])
        self.vars["bunit"].set(self.record["base_unit"])
        self.vars["tunit"].set(self.record.get("trade_unit") or "")
        self.vars["cfact"].set(str(self.record.get("conversion_factor",1.0)))

    def _save(self):
        name  = self.vars["name"].get().strip()
        bunit = self.vars["bunit"].get().strip()
        tunit = self.vars["tunit"].get().strip() or None
        try: cfact = float(self.vars["cfact"].get() or 1.0)
        except ValueError: messagebox.showerror("Error","Conv. factor must be number.", parent=self); return
        if not name or not bunit: messagebox.showerror("Error","Name and Base Unit required.", parent=self); return
        try:
            conn = get_connection()
            if self.record:
                conn.execute("UPDATE products SET name=?,base_unit=?,trade_unit=?,conversion_factor=? WHERE id=?",
                             (name,bunit,tunit,cfact,self.record["id"]))
            else:
                conn.execute("INSERT INTO products (name,base_unit,trade_unit,conversion_factor) VALUES (?,?,?,?)",
                             (name,bunit,tunit,cfact))
            conn.commit(); conn.close()
            if self.on_save: self.on_save()
            self.destroy()
        except Exception as e: messagebox.showerror("Error", str(e), parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  PARTIES
# ═══════════════════════════════════════════════════════════════════════════════
class PartiesFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._build(); self.load_data()

    def _build(self):
        SectionHeader(self, "Parties Management").pack(fill="x")
        bar = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); bar.pack(fill="x")
        StyledButton(bar, "+ Add Party", command=self._add).pack(side="left")
        if PDF_OK: StyledButton(bar, "📄 PDF", command=self._pdf, kind="danger").pack(side="left", padx=8)
        self.table = StyledTable(self, [("id","ID",50),("name","Party Name",250),
                                         ("addr","Address",300),("act","Active",60)])
        # Search bar
        self.search_bar = SearchBar(self, self.table, "🔍  Party name ya address se search karein...")
        self.search_bar.pack(fill="x", padx=PAD, pady=(4,0))
        self.table.pack(fill="both", expand=True, padx=PAD, pady=PAD_SMALL)
        b = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); b.pack(fill="x")
        StyledButton(b, "✏ Edit",   command=self._edit,   kind="warning").pack(side="left", padx=(0,8))
        StyledButton(b, "🗑 Delete", command=self._delete, kind="danger").pack(side="left")

    def load_data(self):
        conn = get_connection()
        rows = conn.execute("SELECT id,name,address,is_active FROM parties ORDER BY name").fetchall()
        conn.close()
        self.table.load([(r["id"],r["name"],r["address"] or "–","Yes" if r["is_active"] else "No") for r in rows])

    def _add(self):    PartyDialog(self, on_save=self.load_data)
    def _edit(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Edit","Select a party."); return
        conn = get_connection()
        row = conn.execute("SELECT * FROM parties WHERE id=?", (sel[0],)).fetchone()
        conn.close()
        if row: PartyDialog(self, record=dict(row), on_save=self.load_data)
    def _delete(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Delete","Select a party."); return
        if not messagebox.askyesno("Delete", f"Delete '{sel[1]}'?", parent=self): return
        conn = get_connection()
        conn.execute("UPDATE parties SET is_active=0 WHERE id=?", (sel[0],))
        conn.commit(); conn.close(); self.load_data()

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        rows = self.table._all_rows
        if not rows: messagebox.showinfo("PDF","Koi party nahi hai."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if not path: return
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        c.setFillColor("#1A2332"); c.rect(0, h-70, w, 70, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h-38, "Saark Industries — Parties List")
        c.setFont("Helvetica", 9)
        c.drawString(40, h-56, f"Total Parties: {len(rows)}   |   Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}")
        y = h - 90
        col_x = [40, 80, 290, 510]
        hdrs  = ["ID", "Party Name", "Address", "Active"]
        c.setFillColor("#2C3E50"); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 9)
        for i, hdr in enumerate(hdrs):
            c.drawString(col_x[i], y+2, hdr)
        y -= 18
        c.setFont("Helvetica", 9)
        for ri, row in enumerate(rows):
            if y < 60: c.showPage(); y = h-60; c.setFont("Helvetica", 9)
            bg = "#F7F9FC" if ri%2==0 else "white"
            c.setFillColor(bg); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
            c.setFillColor("#2C3E50")
            vals = [str(row[0]), str(row[1])[:28], str(row[2])[:30], str(row[3])]
            for i, val in enumerate(vals):
                c.drawString(col_x[i], y+2, val)
            y -= 18
        c.save()
        messagebox.showinfo("PDF", f"PDF saved!\n{path}")


class PartyDialog(ModalDialog):
    def __init__(self, parent, record=None, on_save=None):
        super().__init__(parent, "Add Party" if not record else "Edit Party", 440, 280)
        self.record = record; self.on_save = on_save
        self._build()
        if record: self.name_var.set(record["name"]); self.addr_var.set(record.get("address") or "")

    def _build(self):
        b = self.body; b.columnconfigure(1, weight=1)
        for i, (lbl, var) in enumerate([("Party Name *","name_var"),("Address","addr_var")]):
            tk.Label(b, text=lbl, font=FONT_NORMAL, bg=BG_CARD, fg=TEXT_MAIN
                     ).grid(row=i, column=0, sticky="w", pady=8, padx=(0,12))
            v = tk.StringVar(); setattr(self, var, v)
            tk.Entry(b, textvariable=v, font=FONT_NORMAL, bg=ENTRY_BG, fg=ENTRY_FG,
                     relief="solid", bd=1).grid(row=i, column=1, sticky="ew", ipady=4)
        r = tk.Frame(b, bg=BG_CARD); r.grid(row=2, column=0, columnspan=2, pady=16)
        StyledButton(r, "💾 Save", command=self._save).pack(side="left", padx=(0,8))
        StyledButton(r, "Cancel", command=self.destroy, kind="neutral").pack(side="left")

    def _save(self):
        name = self.name_var.get().strip(); addr = self.addr_var.get().strip() or None
        if not name: messagebox.showerror("Error","Name required.", parent=self); return
        try:
            conn = get_connection()
            if self.record: conn.execute("UPDATE parties SET name=?,address=? WHERE id=?", (name,addr,self.record["id"]))
            else: conn.execute("INSERT INTO parties (name,address) VALUES (?,?)", (name,addr))
            conn.commit(); conn.close()
            if self.on_save: self.on_save()
            self.destroy()
        except Exception as e: messagebox.showerror("Error", str(e), parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  TRANSACTIONS
# ═══════════════════════════════════════════════════════════════════════════════
class TransactionsFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._build(); self.load_data()

    def _build(self):
        SectionHeader(self, "Transactions").pack(fill="x")
        bar = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); bar.pack(fill="x")
        StyledButton(bar, "+ New Transaction", command=self._add).pack(side="left")
        if EXCEL_OK: StyledButton(bar, "⬆ Export Excel", command=self._export, kind="info").pack(side="left", padx=8)
        if PDF_OK:   StyledButton(bar, "📄 PDF",          command=self._pdf,    kind="danger").pack(side="left")

        filt = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=4); filt.pack(fill="x")
        tk.Label(filt, text="From:", font=FONT_SMALL, bg=BG_MAIN, fg=TEXT_MUTED).pack(side="left")
        self.from_var = tk.StringVar(value=today_str())
        make_date_entry(filt, self.from_var).pack(side="left", padx=4)
        tk.Label(filt, text="To:", font=FONT_SMALL, bg=BG_MAIN, fg=TEXT_MUTED).pack(side="left", padx=(8,0))
        self.to_var = tk.StringVar(value=today_str())
        make_date_entry(filt, self.to_var).pack(side="left", padx=4)
        StyledButton(filt, "🔍 Filter", command=self.load_data, kind="info").pack(side="left", padx=8)

        self.table = StyledTable(self, [
            ("id","ID",40),("date","Date",90),("type","Type",100),("party","Party",150),
            ("prod","Product",130),("qty","Qty",70),("unit","Unit",60),("rate","Rate",80),
            ("gst","GST",70),("total","Total",90),("paycash","Pay Cash",85),
            ("payonline","Pay Online",90),("bal","Balance",90),("rem","Remarks",130)])
        # Search bar
        self.search_bar = SearchBar(self, self.table, "🔍  Party, product, type se search karein...")
        self.search_bar.pack(fill="x", padx=PAD, pady=(4,0))
        self.table.pack(fill="both", expand=True, padx=PAD, pady=4)

        b = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=4); b.pack(fill="x")
        StyledButton(b, "✏ Edit",   command=self._edit,   kind="warning").pack(side="left", padx=(0,8))
        StyledButton(b, "🗑 Delete", command=self._delete, kind="danger").pack(side="left")
        tk.Label(b, text="(Edit/Delete: last 7 days only)", font=FONT_SMALL,
                 bg=BG_MAIN, fg=TEXT_MUTED).pack(side="left", padx=12)

    def load_data(self):
        conn = get_connection()
        rows = conn.execute("""
            SELECT t.id,t.date,t.type,p.name as party,pr.name as product,
                   t.entered_quantity,t.entered_unit,t.price_per_unit,t.gst,
                   t.total_amount,t.payment_cash,t.payment_online,t.balance_amount,t.remarks
            FROM transactions t
            JOIN parties p ON p.id=t.party_id JOIN products pr ON pr.id=t.product_id
            WHERE t.date BETWEEN ? AND ? ORDER BY t.date DESC,t.id DESC
        """, (self.from_var.get(), self.to_var.get())).fetchall()
        conn.close()
        self.table.load([(r["id"],r["date"],r["type"],r["party"],r["product"],
                          r["entered_quantity"],r["entered_unit"],f"{r['price_per_unit']:.2f}",
                          f"{r['gst']:.2f}",f"{r['total_amount']:.2f}",f"{r['payment_cash']:.2f}",
                          f"{r['payment_online']:.2f}",f"{r['balance_amount']:.2f}",
                          r["remarks"] or "") for r in rows])

    def _add(self):    TransactionDialog(self, on_save=self.load_data)
    def _edit(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Edit","Select a transaction."); return
        if (datetime.now() - datetime.strptime(sel[1],"%Y-%m-%d")).days > 7:
            messagebox.showwarning("Restricted","Only last 7 days editable."); return
        conn = get_connection()
        row = conn.execute("SELECT * FROM transactions WHERE id=?", (sel[0],)).fetchone()
        conn.close()
        if row: TransactionDialog(self, record=dict(row), on_save=self.load_data)

    def _delete(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Delete","Select a transaction."); return
        if (datetime.now() - datetime.strptime(sel[1],"%Y-%m-%d")).days > 7:
            messagebox.showwarning("Restricted","Only last 7 days deletable."); return
        if not messagebox.askyesno("Delete","Delete and reverse stock?", parent=self): return
        conn = get_connection()
        t = conn.execute("SELECT * FROM transactions WHERE id=?", (sel[0],)).fetchone()
        if t["type"] == "SALE":
            conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (t["quantity_base"],t["product_id"]))
        elif t["type"] == "PURCHASE":
            conn.execute("UPDATE products SET stock=stock-? WHERE id=?", (t["quantity_base"],t["product_id"]))
        # PAYMENT IN / PAYMENT OUT — no stock reversal needed
        conn.execute("DELETE FROM transactions WHERE id=?", (sel[0],))
        conn.commit(); conn.close(); self.load_data()

    def _export(self):
        if not EXCEL_OK: return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        conn = get_connection()
        rows = conn.execute("""
            SELECT t.id,t.date,t.type,p.name,pr.name,t.entered_quantity,t.entered_unit,
                   t.price_per_unit,t.gst,t.total_amount,t.payment_cash,t.payment_online,
                   t.balance_amount,t.remarks
            FROM transactions t JOIN parties p ON p.id=t.party_id JOIN products pr ON pr.id=t.product_id
            ORDER BY t.date DESC""").fetchall()
        conn.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Transactions"
        ws.append(["ID","Date","Type","Party","Product","Qty","Unit","Rate","GST","Total","Pay Cash","Pay Online","Balance","Remarks"])
        for r in rows: ws.append(list(r))
        wb.save(path); messagebox.showinfo("Export", f"Saved: {path}")

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        rows = self.table._all_rows
        if not rows: messagebox.showinfo("PDF","Koi data nahi hai."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if not path: return
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        # Header
        c.setFillColor("#1A2332"); c.rect(0, h-70, w, 70, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h-38, "Saark Industries — Transactions")
        c.setFont("Helvetica", 9)
        c.drawString(40, h-56, f"Period: {self.from_var.get()} to {self.to_var.get()}   |   Total: {len(rows)} records   |   Generated: {datetime.now().strftime('%d %b %Y')}")
        y = h - 90
        # Table headers
        col_x = [40, 80, 145, 215, 300, 355, 385, 415, 445, 480, 520, 560]
        hdrs  = ["ID","Date","Type","Party","Product","Qty","Rate","GST","Total","PayCash","PayOnl","Balance"]
        c.setFillColor("#2C3E50"); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 7)
        for i, hdr in enumerate(hdrs):
            if i < len(col_x): c.drawString(col_x[i], y+2, hdr)
        y -= 18
        c.setFont("Helvetica", 7)
        total_amt = 0.0
        for ri, row in enumerate(rows):
            if y < 60: c.showPage(); y = h-60; c.setFont("Helvetica", 7)
            bg = "#F7F9FC" if ri%2==0 else "white"
            c.setFillColor(bg); c.rect(38, y-2, w-76, 15, fill=1, stroke=0)
            # Color type
            typ = str(row[2]) if len(row) > 2 else ""
            c.setFillColor("#27AE60" if typ=="SALE" else "#2980B9" if typ=="PURCHASE" else "#2C3E50")
            vals = [str(row[0]),str(row[1]),typ,
                    str(row[3])[:12],str(row[4])[:12],
                    str(row[5]),str(row[7]),str(row[8]),
                    str(row[9]),str(row[10]),str(row[11]),str(row[12])]
            for i, val in enumerate(vals):
                if i < len(col_x): c.drawString(col_x[i], y+1, val[:12])
            try: total_amt += float(str(row[9]).replace(",",""))
            except: pass
            y -= 15
        # Footer
        y -= 5
        c.setFillColor("#1A2332"); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 9)
        c.drawString(42, y+2, f"Total Records: {len(rows)}   |   Total Amount: Rs{total_amt:,.2f}")
        c.save()
        messagebox.showinfo("PDF", f"PDF saved!\n{path}")


class TransactionDialog(ModalDialog):
    def __init__(self, parent, record=None, on_save=None):
        super().__init__(parent, "Transaction Entry", 620, 600)
        self.record = record; self.on_save = on_save
        conn = get_connection()
        self.parties  = [dict(r) for r in conn.execute("SELECT id,name FROM parties WHERE is_active=1 ORDER BY name").fetchall()]
        self.products = [dict(r) for r in conn.execute("SELECT id,name,base_unit,trade_unit,conversion_factor FROM products WHERE is_active=1 ORDER BY name").fetchall()]
        conn.close()
        self._build()
        if record: self._populate()
        else: self.date_var.set(today_str())

    def _build(self):
        b = self.body; b.columnconfigure(1, weight=1); b.columnconfigure(3, weight=1)

        # Date & Type
        tk.Label(b,text="Date *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=0,column=0,sticky="w",pady=4)
        self.date_var = tk.StringVar()
        make_date_entry(b, self.date_var).grid(row=0,column=1,sticky="w",padx=(4,16))
        tk.Label(b,text="Type *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=0,column=2,sticky="w",pady=4)
        self.type_var = tk.StringVar(value="SALE")
        type_cb = ttk.Combobox(b,textvariable=self.type_var,values=["SALE","PURCHASE","PAYMENT IN","PAYMENT OUT"],state="readonly",width=14,font=FONT_NORMAL)
        type_cb.grid(row=0,column=3,sticky="w",padx=4)
        type_cb.bind("<<ComboboxSelected>>", self._on_type)

        # Party & Product
        tk.Label(b,text="Party *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=1,column=0,sticky="w",pady=4)
        self.party_var = tk.StringVar()
        ttk.Combobox(b,textvariable=self.party_var,values=[p["name"] for p in self.parties],state="readonly",width=20,font=FONT_NORMAL).grid(row=1,column=1,sticky="ew",padx=(4,16))
        self.prod_lbl = tk.Label(b,text="Product",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED)
        self.prod_lbl.grid(row=1,column=2,sticky="w",pady=4)
        self.prod_var = tk.StringVar()
        self.prod_cb = ttk.Combobox(b,textvariable=self.prod_var,values=[p["name"] for p in self.products],state="readonly",width=20,font=FONT_NORMAL)
        self.prod_cb.grid(row=1,column=3,sticky="ew",padx=4)
        self.prod_cb.bind("<<ComboboxSelected>>", self._on_prod)

        # Qty & Unit
        self.qty_lbl = tk.Label(b,text="Quantity",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED)
        self.qty_lbl.grid(row=2,column=0,sticky="w",pady=4)
        self.qty_var = tk.StringVar(value="0")
        self.qty_entry = tk.Entry(b,textvariable=self.qty_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=14)
        self.qty_entry.grid(row=2,column=1,sticky="w",padx=(4,16),ipady=4)
        tk.Label(b,text="Unit",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=2,column=2,sticky="w",pady=4)
        self.unit_var = tk.StringVar()
        self.unit_cb = ttk.Combobox(b,textvariable=self.unit_var,values=[],state="readonly",width=12)
        self.unit_cb.grid(row=2,column=3,sticky="w",padx=4)

        # Rate & GST
        self.rate_lbl = tk.Label(b,text="Rate",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED)
        self.rate_lbl.grid(row=3,column=0,sticky="w",pady=4)
        self.rate_var = tk.StringVar(value="0")
        self.rate_entry = tk.Entry(b,textvariable=self.rate_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=14)
        self.rate_entry.grid(row=3,column=1,sticky="w",padx=(4,16),ipady=4)

        # GST % + Amt in same row
        gst_frame = tk.Frame(b, bg=BG_CARD)
        gst_frame.grid(row=3,column=2,columnspan=2,sticky="w",padx=4)
        tk.Label(gst_frame,text="GST %",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).pack(side="left")
        self.gst_pct_var = tk.StringVar(value="0")
        tk.Entry(gst_frame,textvariable=self.gst_pct_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=6).pack(side="left",padx=(4,8),ipady=4)
        tk.Label(gst_frame,text="GST Amt",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).pack(side="left")
        self.gst_var = tk.StringVar(value="0")
        tk.Entry(gst_frame,textvariable=self.gst_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=8).pack(side="left",padx=4,ipady=4)

        # Amount field (for PAYMENT IN/OUT)
        self.amt_lbl = tk.Label(b,text="Amount *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED)
        self.amt_lbl.grid(row=4,column=0,sticky="w",pady=4)
        self.amt_var = tk.StringVar(value="0")
        self.amt_entry = tk.Entry(b,textvariable=self.amt_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=14)
        self.amt_entry.grid(row=4,column=1,sticky="w",padx=(4,16),ipady=4)

        # Total & Pay Cash
        tk.Label(b,text="Total",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=4,column=2,sticky="w",pady=4)
        self.total_var = tk.StringVar(value="0.00")
        tk.Entry(b,textvariable=self.total_var,font=FONT_BOLD,bg="#EBF5EB",fg=ACCENT_DARK,relief="solid",bd=1,width=14,state="readonly").grid(row=4,column=3,sticky="w",padx=4,ipady=4)

        # Pay Cash & Pay Online
        tk.Label(b,text="Pay Cash",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=5,column=0,sticky="w",pady=4)
        self.advc_var = tk.StringVar(value="0")
        tk.Entry(b,textvariable=self.advc_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=14).grid(row=5,column=1,sticky="w",padx=(4,16),ipady=4)
        tk.Label(b,text="Pay Online",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=5,column=2,sticky="w",pady=4)
        self.advo_var = tk.StringVar(value="0")
        tk.Entry(b,textvariable=self.advo_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=14).grid(row=5,column=3,sticky="w",padx=4,ipady=4)

        # Balance + Stock Info
        tk.Label(b,text="Balance",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=6,column=0,sticky="w",pady=4)
        self.bal_var = tk.StringVar(value="0.00")
        tk.Entry(b,textvariable=self.bal_var,font=FONT_BOLD,bg="#FEF9E7",fg=WARNING_DARK,relief="solid",bd=1,width=14,state="readonly").grid(row=6,column=1,sticky="w",padx=(4,16),ipady=4)
        tk.Label(b,text="Stock Baad Mein",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=6,column=2,sticky="w",pady=4)
        self.stock_after_var = tk.StringVar(value="—")
        tk.Entry(b,textvariable=self.stock_after_var,font=FONT_BOLD,bg="#EBF5FB",fg=INFO,relief="solid",bd=1,width=14,state="readonly").grid(row=6,column=3,sticky="w",padx=4,ipady=4)

        # Remarks
        tk.Label(b,text="Remarks",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=7,column=0,sticky="w",pady=4)
        self.rem_var = tk.StringVar()
        tk.Entry(b,textvariable=self.rem_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1).grid(row=7,column=1,columnspan=3,sticky="ew",padx=4,ipady=4)

        # Trace for auto-calc
        for v in [self.qty_var, self.rate_var, self.gst_var, self.gst_pct_var, self.advc_var, self.advo_var, self.amt_var]:
            v.trace_add("write", self._recalc)

        r = tk.Frame(b, bg=BG_CARD); r.grid(row=8, column=0, columnspan=4, pady=14)
        StyledButton(r, "💾 Save", command=self._save).pack(side="left", padx=(0,8))
        StyledButton(r, "Cancel", command=self.destroy, kind="neutral").pack(side="left")

        # Set default view
        self._on_type()

    def _on_type(self, _=None):
        """Show/hide fields based on transaction type."""
        txn_type = self.type_var.get()
        is_payment = txn_type in ("PAYMENT IN", "PAYMENT OUT")
        state_sale = "disabled" if is_payment else "normal"
        self.prod_cb.config(state="disabled" if is_payment else "readonly")
        self.qty_entry.config(state=state_sale)
        self.rate_entry.config(state=state_sale)
        self.unit_cb.config(state="disabled" if is_payment else "readonly")
        self.amt_entry.config(state="normal" if is_payment else "disabled")
        if is_payment:
            self.qty_var.set("0"); self.rate_var.set("0")
            self.gst_var.set("0"); self.gst_pct_var.set("0")
            self.prod_var.set(""); self.unit_var.set("")
            self.stock_after_var.set("—")
            self.amt_lbl.config(fg=TEXT_MAIN)
        else:
            self.amt_var.set("0")
            self.amt_lbl.config(fg=TEXT_MUTED)
        self._recalc()

    def _on_prod(self, _=None):
        prod = next((p for p in self.products if p["name"] == self.prod_var.get()), None)
        if prod:
            units = [prod["base_unit"]] + ([prod["trade_unit"]] if prod["trade_unit"] else [])
            self.unit_cb["values"] = units; self.unit_var.set(units[0])
        self._recalc()

    def _recalc(self, *_):
        try:
            txn_type   = self.type_var.get()
            is_payment = txn_type in ("PAYMENT IN", "PAYMENT OUT")
            advc = float(self.advc_var.get() or 0)
            advo = float(self.advo_var.get() or 0)
            if is_payment:
                total = float(self.amt_var.get() or 0)
                self.stock_after_var.set("—")
            else:
                qty      = float(self.qty_var.get() or 0)
                rate     = float(self.rate_var.get() or 0)
                gst_pct  = float(self.gst_pct_var.get() or 0)
                if gst_pct > 0:
                    gst_amt = round(qty * rate * gst_pct / 100, 2)
                    self.gst_var.set(str(gst_amt))
                gst   = float(self.gst_var.get() or 0)
                total = qty * rate + gst
                # Stock after calculation
                prod_n = self.prod_var.get()
                prod   = next((p for p in self.products if p["name"] == prod_n), None)
                if prod and qty > 0:
                    unit     = self.unit_var.get()
                    qty_base = qty * prod["conversion_factor"] if prod.get("trade_unit") and unit == prod["trade_unit"] else qty
                    conn = get_connection()
                    cur_stock = conn.execute("SELECT stock FROM products WHERE id=?", (prod["id"],)).fetchone()
                    conn.close()
                    if cur_stock:
                        st = cur_stock["stock"]
                        if txn_type == "SALE":
                            after = st - qty_base
                            color = DANGER if after < 0 else INFO
                        else:
                            after = st + qty_base
                            color = ACCENT
                        self.stock_after_var.set(f"{after:.2f} {prod['base_unit']}")
                else:
                    self.stock_after_var.set("—")
            self.total_var.set(f"{total:.2f}")
            self.bal_var.set(f"{total - advc - advo:.2f}")
        except ValueError: pass

    def _populate(self):
        r = self.record; self.date_var.set(r["date"]); self.type_var.set(r["type"])
        self._on_type()
        conn = get_connection()
        party   = conn.execute("SELECT name FROM parties  WHERE id=?", (r["party_id"],)).fetchone()
        product = conn.execute("SELECT * FROM products WHERE id=?", (r["product_id"],)).fetchone()
        conn.close()
        if party: self.party_var.set(party["name"])
        txn_type = r["type"]
        is_payment = txn_type in ("PAYMENT IN", "PAYMENT OUT")
        if not is_payment:
            if product: self.prod_var.set(product["name"]); self._on_prod()
            self.qty_var.set(str(r["entered_quantity"])); self.unit_var.set(r["entered_unit"])
            self.rate_var.set(str(r["price_per_unit"])); self.gst_var.set(str(r["gst"]))
        else:
            self.amt_var.set(str(r["total_amount"]))
        self.advc_var.set(str(r["payment_cash"])); self.advo_var.set(str(r["payment_online"]))
        self.rem_var.set(r.get("remarks") or "")

    def _save(self):
        date     = self.date_var.get().strip()
        txn_type = self.type_var.get()
        party_n  = self.party_var.get()
        is_payment = txn_type in ("PAYMENT IN", "PAYMENT OUT")

        if not date or not party_n:
            messagebox.showerror("Error","Date aur Party zaroori hai.", parent=self); return
        party = next((p for p in self.parties if p["name"] == party_n), None)
        if not party:
            messagebox.showerror("Error","Party select karein.", parent=self); return

        try:
            advc = float(self.advc_var.get() or 0)
            advo = float(self.advo_var.get() or 0)
        except ValueError:
            messagebox.showerror("Error","Pay Cash/Online numbers hone chahiye.", parent=self); return

        remarks = self.rem_var.get().strip() or None

        # ── PAYMENT IN / PAYMENT OUT ──────────────────────────────────────────
        if is_payment:
            try: total = float(self.amt_var.get() or 0)
            except ValueError:
                messagebox.showerror("Error","Amount number hona chahiye.", parent=self); return
            if total <= 0:
                messagebox.showerror("Error","Amount 0 se zyada hona chahiye.", parent=self); return
            balance = total - advc - advo
            # Use a dummy product (first available) for DB constraint
            conn = get_connection()
            dummy = conn.execute("SELECT id FROM products LIMIT 1").fetchone()
            if not dummy:
                messagebox.showerror("Error","Pehle ek product add karein.", parent=self)
                conn.close(); return
            prod_id = dummy["id"]
            try:
                if self.record:
                    conn.execute("""UPDATE transactions SET date=?,type=?,party_id=?,product_id=?,
                        quantity_base=0,entered_quantity=0,entered_unit='-',price_per_unit=0,
                        gst=0,total_amount=?,payment_cash=?,payment_online=?,balance_amount=?,remarks=?
                        WHERE id=?""",
                        (date,txn_type,party["id"],prod_id,total,advc,advo,balance,remarks,self.record["id"]))
                else:
                    conn.execute("""INSERT INTO transactions (date,type,party_id,product_id,
                        quantity_base,entered_quantity,entered_unit,price_per_unit,gst,
                        total_amount,payment_cash,payment_online,balance_amount,remarks)
                        VALUES (?,?,?,?,0,0,'-',0,0,?,?,?,?,?)""",
                        (date,txn_type,party["id"],prod_id,total,advc,advo,balance,remarks))
                conn.commit(); conn.close()
                if self.on_save: self.on_save()
                self.destroy()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=self)
            return

        # ── SALE / PURCHASE ───────────────────────────────────────────────────
        prod_n = self.prod_var.get()
        unit   = self.unit_var.get()
        if not prod_n or not unit:
            messagebox.showerror("Error","Product aur Unit zaroori hai.", parent=self); return
        product = next((p for p in self.products if p["name"] == prod_n), None)
        if not product:
            messagebox.showerror("Error","Product select karein.", parent=self); return
        try:
            qty  = float(self.qty_var.get() or 0)
            rate = float(self.rate_var.get() or 0)
            gst  = float(self.gst_var.get() or 0)
        except ValueError:
            messagebox.showerror("Error","Qty/Rate numbers hone chahiye.", parent=self); return
        if qty <= 0:
            messagebox.showerror("Error","Quantity 0 se zyada honi chahiye.", parent=self); return

        qty_base = qty * product["conversion_factor"] if product["trade_unit"] and unit == product["trade_unit"] else qty
        total    = qty * rate + gst
        balance  = total - advc - advo

        try:
            conn = get_connection()
            if self.record:
                old = conn.execute("SELECT * FROM transactions WHERE id=?", (self.record["id"],)).fetchone()
                if old["type"] == "SALE":
                    conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (old["quantity_base"], old["product_id"]))
                elif old["type"] == "PURCHASE":
                    conn.execute("UPDATE products SET stock=stock-? WHERE id=?", (old["quantity_base"], old["product_id"]))
            if txn_type == "SALE":
                cur_stock = conn.execute("SELECT stock FROM products WHERE id=?", (product["id"],)).fetchone()["stock"]
                if cur_stock < qty_base:
                    messagebox.showerror("Stock Error", f"Stock kam hai!\nAvailable: {cur_stock:.2f}", parent=self)
                    conn.close(); return
                conn.execute("UPDATE products SET stock=stock-? WHERE id=?", (qty_base, product["id"]))
            elif txn_type == "PURCHASE":
                conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (qty_base, product["id"]))
            if self.record:
                conn.execute("""UPDATE transactions SET date=?,type=?,party_id=?,product_id=?,quantity_base=?,
                    entered_quantity=?,entered_unit=?,price_per_unit=?,gst=?,total_amount=?,payment_cash=?,
                    payment_online=?,balance_amount=?,remarks=? WHERE id=?""",
                    (date,txn_type,party["id"],product["id"],qty_base,qty,unit,rate,gst,total,advc,advo,balance,remarks,self.record["id"]))
            else:
                conn.execute("""INSERT INTO transactions (date,type,party_id,product_id,quantity_base,
                    entered_quantity,entered_unit,price_per_unit,gst,total_amount,payment_cash,
                    payment_online,balance_amount,remarks) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (date,txn_type,party["id"],product["id"],qty_base,qty,unit,rate,gst,total,advc,advo,balance,remarks))
            conn.commit(); conn.close()
            if self.on_save: self.on_save()
            self.destroy()
        except Exception as e: messagebox.showerror("Error", str(e), parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  PRODUCTION
# ═══════════════════════════════════════════════════════════════════════════════
class MaterialRow(tk.Frame):
    def __init__(self, parent, products, **kw):
        super().__init__(parent, bg=BG_CARD, **kw)
        self.products = products
        self.prod_var = tk.StringVar(); self.qty_var = tk.StringVar(); self.unit_var = tk.StringVar()
        pc = ttk.Combobox(self, textvariable=self.prod_var, values=[p["name"] for p in products],
                          state="readonly", width=20, font=FONT_SMALL)
        pc.pack(side="left", padx=2); pc.bind("<<ComboboxSelected>>", self._on_prod)
        tk.Entry(self, textvariable=self.qty_var, width=10, font=FONT_SMALL,
                 bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1).pack(side="left", padx=2, ipady=3)
        self.unit_cb = ttk.Combobox(self, textvariable=self.unit_var, values=[], state="readonly", width=10)
        self.unit_cb.pack(side="left", padx=2)
        self.remove_btn = StyledButton(self, "✕", kind="danger"); self.remove_btn.pack(side="left", padx=4)

    def _on_prod(self, _=None):
        prod = next((p for p in self.products if p["name"] == self.prod_var.get()), None)
        if prod:
            units = [prod["base_unit"]] + ([prod["trade_unit"]] if prod["trade_unit"] else [])
            self.unit_cb["values"] = units; self.unit_var.set(units[0])

    def get_data(self): return self.prod_var.get(), self.qty_var.get(), self.unit_var.get()


class ProductionFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self.input_rows = []; self.output_rows = []
        self._load_products(); self._build()

    def _load_products(self):
        conn = get_connection()
        self.products = [dict(r) for r in conn.execute(
            "SELECT id,name,base_unit,trade_unit,conversion_factor,stock FROM products WHERE is_active=1").fetchall()]
        conn.close()

    def _build(self):
        SectionHeader(self, "Production Entry").pack(fill="x")
        top = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=PAD_SMALL); top.pack(fill="x")
        tk.Label(top, text="Production Date:", font=FONT_BOLD, bg=BG_MAIN, fg=TEXT_MAIN).pack(side="left")
        self.date_var = tk.StringVar(value=today_str())
        make_date_entry(top, self.date_var).pack(side="left", padx=8)
        StyledButton(top, "💾 Save Batch", command=self._save).pack(side="right")

        panels = tk.Frame(self, bg=BG_MAIN); panels.pack(fill="both", expand=True, padx=PAD, pady=PAD_SMALL)
        panels.columnconfigure(0, weight=1); panels.columnconfigure(1, weight=1)

        left = tk.LabelFrame(panels, text=" ▼ Input Materials (Stock Reduce) ",
                             font=FONT_BOLD, bg=BG_CARD, fg=DANGER, padx=8, pady=8)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        self._inp_container = tk.Frame(left, bg=BG_CARD); self._inp_container.pack(fill="both", expand=True)
        StyledButton(left, "+ Add Input", command=lambda: self._add_row("input"), kind="danger").pack(pady=6)

        right = tk.LabelFrame(panels, text=" ▲ Output Materials (Stock Increase) ",
                              font=FONT_BOLD, bg=BG_CARD, fg=ACCENT, padx=8, pady=8)
        right.grid(row=0, column=1, sticky="nsew", padx=(6,0))
        self._out_container = tk.Frame(right, bg=BG_CARD); self._out_container.pack(fill="both", expand=True)
        StyledButton(right, "+ Add Output", command=lambda: self._add_row("output")).pack(pady=6)

    def _add_row(self, kind):
        self._load_products()
        container = self._inp_container if kind == "input" else self._out_container
        rows_list  = self.input_rows if kind == "input" else self.output_rows
        row = MaterialRow(container, self.products)
        row.pack(fill="x", pady=2)
        rows_list.append(row)
        row.remove_btn.config(command=lambda r=row, rl=rows_list: (r.destroy(), rl.remove(r)) if r in rl else None)

    def _save(self):
        self._load_products()
        if not self.input_rows and not self.output_rows:
            messagebox.showwarning("Empty","Add at least one row."); return

        def parse(rows):
            items = []
            for row in rows:
                pn, qs, unit = row.get_data()
                if not pn: messagebox.showerror("Error","Select product for all rows.", parent=self); return None
                try: qty = float(qs)
                except ValueError: messagebox.showerror("Error",f"Invalid qty for {pn}.", parent=self); return None
                prod = next((p for p in self.products if p["name"] == pn), None)
                if not prod: return None
                qty_base = qty * prod["conversion_factor"] if prod["trade_unit"] and unit == prod["trade_unit"] else qty
                items.append({"product":prod,"qty":qty,"unit":unit,"qty_base":qty_base})
            return items

        inputs = parse(self.input_rows); outputs = parse(self.output_rows)
        if inputs is None or outputs is None: return

        conn = get_connection()
        for item in inputs:
            cur = conn.execute("SELECT stock FROM products WHERE id=?", (item["product"]["id"],)).fetchone()["stock"]
            if cur < item["qty_base"]:
                messagebox.showerror("Stock Error",
                    f"Insufficient stock for '{item['product']['name']}'!\nAvailable: {cur:.2f}, Required: {item['qty_base']:.2f}")
                conn.close(); return

        cur = conn.execute("INSERT INTO production_batches (date) VALUES (?)", (self.date_var.get(),))
        batch_id = cur.lastrowid
        for item in inputs:
            conn.execute("INSERT INTO production_inputs (batch_id,product_id,quantity_base,entered_quantity,entered_unit) VALUES (?,?,?,?,?)",
                         (batch_id,item["product"]["id"],item["qty_base"],item["qty"],item["unit"]))
            conn.execute("UPDATE products SET stock=stock-? WHERE id=?", (item["qty_base"],item["product"]["id"]))
        for item in outputs:
            conn.execute("INSERT INTO production_outputs (batch_id,product_id,quantity_base,entered_quantity,entered_unit) VALUES (?,?,?,?,?)",
                         (batch_id,item["product"]["id"],item["qty_base"],item["qty"],item["unit"]))
            conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (item["qty_base"],item["product"]["id"]))
        conn.commit(); conn.close()
        messagebox.showinfo("Success", f"Batch #{batch_id} saved!")
        for r in self.input_rows[:]:  r.destroy()
        for r in self.output_rows[:]: r.destroy()
        self.input_rows.clear(); self.output_rows.clear()


# ═══════════════════════════════════════════════════════════════════════════════
#  PARTY LEDGER
# ═══════════════════════════════════════════════════════════════════════════════
class PartyLedgerFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._rows_cache = []; self._build()

    def _build(self):
        SectionHeader(self, "Party Ledger").pack(fill="x")
        conn = get_connection()
        self._parties  = [dict(r) for r in conn.execute("SELECT id,name FROM parties ORDER BY name").fetchall()]
        self._products = [dict(r) for r in conn.execute("SELECT id,name FROM products ORDER BY name").fetchall()]
        conn.close()

        filt = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=PAD_SMALL); filt.pack(fill="x")
        tk.Label(filt,text="Party:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).grid(row=0,column=0,padx=(0,4))
        self.party_var = tk.StringVar(value="All")
        ttk.Combobox(filt,textvariable=self.party_var,values=["All"]+[p["name"] for p in self._parties],state="readonly",width=22,font=FONT_NORMAL).grid(row=0,column=1,padx=4)
        tk.Label(filt,text="Product:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).grid(row=0,column=2,padx=(12,4))
        self.prod_var = tk.StringVar(value="All")
        ttk.Combobox(filt,textvariable=self.prod_var,values=["All"]+[p["name"] for p in self._products],state="readonly",width=22,font=FONT_NORMAL).grid(row=0,column=3,padx=4)
        tk.Label(filt,text="From:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).grid(row=0,column=4,padx=(12,4))
        self.from_var = tk.StringVar(value="2020-01-01")
        make_date_entry(filt, self.from_var).grid(row=0,column=5,padx=4)
        tk.Label(filt,text="To:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).grid(row=0,column=6,padx=(8,4))
        self.to_var = tk.StringVar(value=today_str())
        make_date_entry(filt, self.to_var).grid(row=0,column=7,padx=4)
        StyledButton(filt,"📊 Generate",command=self._generate,kind="info").grid(row=0,column=8,padx=(12,0))
        if EXCEL_OK: StyledButton(filt,"⬆ Excel",command=self._export,kind="neutral").grid(row=0,column=9,padx=4)
        if PDF_OK:   StyledButton(filt,"📄 PDF",  command=self._pdf,   kind="danger" ).grid(row=0,column=10,padx=4)

        self.outstanding_lbl = tk.Label(self, text="Outstanding: ₹0.00", font=FONT_SUBTITLE, bg=BG_MAIN, fg=ACCENT)
        self.outstanding_lbl.pack(anchor="e", padx=PAD)
        self.table = StyledTable(self, [
            ("date","Date",90),("type","Type",95),("prod","Product",140),
            ("qty","Qty",65),("rate","Rate",80),
            ("debit","Debit (Dr)",100),("credit","Credit (Cr)",100),
            ("paycash","Pay Cash",85),("payonline","Pay Online",85),
            ("bal","Balance",100)])
        # Search bar
        self.search_bar = SearchBar(self, self.table, "🔍  Type, product se search karein...")
        self.search_bar.pack(fill="x", padx=PAD, pady=(4,0))
        self.table.pack(fill="both", expand=True, padx=PAD, pady=PAD_SMALL)

    def _generate(self):
        filters = ["t.date BETWEEN ? AND ?"]; params = [self.from_var.get(), self.to_var.get()]
        if self.party_var.get() != "All":
            pid = next((p["id"] for p in self._parties if p["name"] == self.party_var.get()), None)
            if pid: filters.append("t.party_id=?"); params.append(pid)
        if self.prod_var.get() != "All":
            prid = next((p["id"] for p in self._products if p["name"] == self.prod_var.get()), None)
            if prid: filters.append("t.product_id=?"); params.append(prid)
        conn = get_connection()
        rows = conn.execute(f"""
            SELECT t.date, t.type, pr.name as product,
                   t.entered_quantity, t.price_per_unit, t.total_amount,
                   t.payment_cash, t.payment_online, t.balance_amount
            FROM transactions t JOIN products pr ON pr.id=t.product_id
            WHERE {' AND '.join(filters)} ORDER BY t.date, t.id""", params).fetchall()
        conn.close()
        data = []; running = 0.0
        for r in rows:
            typ = r["type"]
            amt = r["total_amount"]
            pc  = r["payment_cash"]
            po  = r["payment_online"]
            if typ == "SALE":
                debit = amt; credit = 0.0; running += amt
                # payments received reduce outstanding
                running -= (pc + po)
            elif typ == "PURCHASE":
                credit = amt; debit = 0.0; running -= amt
                running += (pc + po)
            elif typ == "PAYMENT IN":
                credit = amt; debit = 0.0; running -= amt
            elif typ == "PAYMENT OUT":
                debit = amt; credit = 0.0; running += amt
            else:
                debit = 0.0; credit = 0.0
            data.append((r["date"], typ, r["product"],
                         f"{r['entered_quantity']:.2f}", f"{r['price_per_unit']:.2f}",
                         f"{debit:.2f}", f"{credit:.2f}",
                         f"{pc:.2f}", f"{po:.2f}", f"{running:.2f}"))
        self.table.load(data); self._rows_cache = data
        self.outstanding_lbl.config(text=f"Outstanding: ₹{running:.2f}", fg=DANGER if running > 0 else ACCENT)

    def _export(self):
        if not EXCEL_OK or not self._rows_cache: return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Party Ledger"
        ws.append(["Date","Type","Product","Qty","Rate","Debit","Credit","Pay Cash","Pay Online","Balance"])
        for r in self._rows_cache: ws.append(list(r))
        wb.save(path); messagebox.showinfo("Export", f"Saved: {path}")

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        if not self._rows_cache: messagebox.showinfo("PDF","Pehle Generate karein."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if not path: return
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        # Header
        c.setFillColor("#1A2332"); c.rect(0, h-70, w, 70, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h-38, "Saark Industries — Party Ledger")
        c.setFont("Helvetica", 9)
        party_txt = self.party_var.get()
        c.drawString(40, h-56, f"Party: {party_txt}   |   Period: {self.from_var.get()} to {self.to_var.get()}   |   Generated: {datetime.now().strftime('%d %b %Y')}")
        # Outstanding
        outstanding_txt = self.outstanding_lbl.cget("text")
        c.setFont("Helvetica-Bold", 11)
        c.drawString(w-200, h-56, outstanding_txt)
        y = h - 90
        # Table headers
        col_x = [40, 100, 170, 285, 330, 375, 430, 480, 520, 565]
        hdrs  = ["Date","Type","Product","Qty","Rate","Debit","Credit","PayCash","PayOnl","Balance"]
        c.setFillColor("#2C3E50"); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 7)
        for i, hdr in enumerate(hdrs):
            if i < len(col_x): c.drawString(col_x[i], y+2, hdr)
        y -= 18
        c.setFont("Helvetica", 7)
        for ri, row in enumerate(self._rows_cache):
            if y < 60: c.showPage(); y = h-60; c.setFont("Helvetica", 7)
            bg = "#F7F9FC" if ri%2==0 else "white"
            c.setFillColor(bg); c.rect(38, y-2, w-76, 15, fill=1, stroke=0)
            c.setFillColor("#2C3E50")
            for i, val in enumerate(row):
                if i < len(col_x): c.drawString(col_x[i], y+1, str(val)[:14])
            y -= 15
        # Footer outstanding
        y -= 5
        c.setFillColor("#1A2332"); c.rect(38, y-2, w-76, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 9)
        c.drawString(42, y+2, outstanding_txt)
        c.save()
        messagebox.showinfo("PDF", f"PDF saved!\n{path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  PROFIT / LOSS
# ═══════════════════════════════════════════════════════════════════════════════
class ProfitLossFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._data = {}; self._build()

    def _build(self):
        SectionHeader(self, "Profit / Loss Report").pack(fill="x")
        filt = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=PAD_SMALL); filt.pack(fill="x")
        tk.Label(filt,text="From:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left")
        self.from_var = tk.StringVar(value="2020-01-01")
        make_date_entry(filt, self.from_var).pack(side="left", padx=4)
        tk.Label(filt,text="To:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left", padx=(8,0))
        self.to_var = tk.StringVar(value=today_str())
        make_date_entry(filt, self.to_var).pack(side="left", padx=4)
        StyledButton(filt,"📊 Generate",command=self._generate,kind="info").pack(side="left", padx=8)
        if EXCEL_OK: StyledButton(filt,"⬆ Excel",command=self._export,kind="neutral").pack(side="left", padx=(0,6))
        if PDF_OK:   StyledButton(filt,"📄 PDF",  command=self._pdf,   kind="danger" ).pack(side="left")

        cards = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=PAD); cards.pack(fill="x")
        self._cards = {}
        for col, (title, color, key) in enumerate([
            ("Total Sales",ACCENT,"sales"),("Total Purchases",INFO,"purchases"),
            ("Total Expenses",WARNING,"expenses"),("Net Profit/Loss",DANGER,"net")]):
            f = tk.Frame(cards, bg=color, padx=20, pady=16, cursor="hand2")
            f.grid(row=0, column=col, padx=8, sticky="ew"); cards.columnconfigure(col, weight=1)
            tk.Label(f, text=title, font=FONT_SMALL, bg=color, fg="white").pack()
            lbl = tk.Label(f, text="₹0.00", font=(FONT_FAMILY,18,"bold"), bg=color, fg="white"); lbl.pack()
            if key != "net":
                tk.Label(f, text="(click for details)", font=(FONT_FAMILY,7), bg=color, fg="white").pack()
                f.bind("<Button-1>", lambda e, k=key: self._show_details(k))
                lbl.bind("<Button-1>", lambda e, k=key: self._show_details(k))
            self._cards[key] = lbl

        # Detail table below cards
        self.detail_lbl = tk.Label(self, text="", font=FONT_BOLD, bg=BG_MAIN, fg=TEXT_MAIN, padx=PAD)
        self.detail_lbl.pack(anchor="w", pady=(8,0))
        self.detail_table = StyledTable(self, [
            ("date","Date",90),("party","Party",160),("product","Product",150),
            ("qty","Qty",80),("rate","Rate",90),("gst","GST",80),("total","Amount",110)])
        self.detail_table.pack(fill="both", expand=True, padx=PAD, pady=4)

    def _generate(self):
        fd = self.from_var.get(); td = self.to_var.get()
        conn = get_connection()
        sales     = conn.execute("SELECT COALESCE(SUM(total_amount),0) FROM transactions WHERE type='SALE' AND date BETWEEN ? AND ?", (fd,td)).fetchone()[0]
        purchases = conn.execute("SELECT COALESCE(SUM(total_amount),0) FROM transactions WHERE type='PURCHASE' AND date BETWEEN ? AND ?", (fd,td)).fetchone()[0]
        expenses  = conn.execute("SELECT COALESCE(SUM(amount),0) FROM expenses WHERE date BETWEEN ? AND ?", (fd,td)).fetchone()[0]
        conn.close()
        net = sales - purchases - expenses
        self._data = {"sales":sales,"purchases":purchases,"expenses":expenses,"net":net}
        self._cards["sales"].config(text=f"₹{sales:,.2f}")
        self._cards["purchases"].config(text=f"₹{purchases:,.2f}")
        self._cards["expenses"].config(text=f"₹{expenses:,.2f}")
        nc = ACCENT if net >= 0 else DANGER
        self._cards["net"].config(text=f"₹{net:,.2f}")
        self._cards["net"].master.config(bg=nc); self._cards["net"].config(bg=nc)
        self.detail_table.load([])
        self.detail_lbl.config(text="")

    def _show_details(self, key):
        if not self._data: messagebox.showinfo("Info","Pehle Generate karein."); return
        fd = self.from_var.get(); td = self.to_var.get()
        conn = get_connection()
        if key == "sales":
            self.detail_lbl.config(text="📋 Sales Details:", fg=ACCENT)
            rows = conn.execute("""
                SELECT t.date, p.name as party, pr.name as product,
                       t.entered_quantity, t.price_per_unit, t.gst, t.total_amount
                FROM transactions t
                JOIN parties p ON p.id=t.party_id JOIN products pr ON pr.id=t.product_id
                WHERE t.type='SALE' AND t.date BETWEEN ? AND ? ORDER BY t.date""", (fd,td)).fetchall()
            data = [(r["date"],r["party"],r["product"],f"{r['entered_quantity']:.2f}",
                     f"{r['price_per_unit']:.2f}",f"{r['gst']:.2f}",f"{r['total_amount']:.2f}") for r in rows]
        elif key == "purchases":
            self.detail_lbl.config(text="📋 Purchase Details:", fg=INFO)
            rows = conn.execute("""
                SELECT t.date, p.name as party, pr.name as product,
                       t.entered_quantity, t.price_per_unit, t.gst, t.total_amount
                FROM transactions t
                JOIN parties p ON p.id=t.party_id JOIN products pr ON pr.id=t.product_id
                WHERE t.type='PURCHASE' AND t.date BETWEEN ? AND ? ORDER BY t.date""", (fd,td)).fetchall()
            data = [(r["date"],r["party"],r["product"],f"{r['entered_quantity']:.2f}",
                     f"{r['price_per_unit']:.2f}",f"{r['gst']:.2f}",f"{r['total_amount']:.2f}") for r in rows]
        elif key == "expenses":
            self.detail_lbl.config(text="📋 Expense Details:", fg=WARNING)
            rows = conn.execute("""
                SELECT e.date, COALESCE(emp.name,'-') as party, e.category as product,
                       1 as qty, e.amount as rate, 0 as gst, e.amount
                FROM expenses e LEFT JOIN employees emp ON emp.id=e.employee_id
                WHERE e.date BETWEEN ? AND ? ORDER BY e.date""", (fd,td)).fetchall()
            data = [(r["date"],r["party"],r["product"],"—",
                     "—","—",f"{r['amount']:.2f}") for r in rows]
        else:
            data = []
        conn.close()
        self.detail_table.load(data)

    def _export(self):
        if not EXCEL_OK or not self._data: messagebox.showinfo("Export","Generate first."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "P&L"
        ws.append(["Metric","Amount"])
        for k,v in [("Total Sales",self._data["sales"]),("Total Purchases",self._data["purchases"]),
                    ("Total Expenses",self._data["expenses"]),("Net Profit/Loss",self._data["net"])]:
            ws.append([k,v])
        wb.save(path); messagebox.showinfo("Export", f"Saved: {path}")

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        if not self._data: messagebox.showinfo("PDF","Pehle Generate karein."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if not path: return
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        # Header
        c.setFillColor("#2ECC71"); c.rect(0, h-70, w, 70, fill=1, stroke=0)
        c.setFillColor("white")
        c.setFont("Helvetica-Bold", 18)
        c.drawString(40, h-40, "Saark Industries — Profit / Loss Report")
        c.setFont("Helvetica", 10)
        fd = self.from_var.get(); td = self.to_var.get()
        c.drawString(40, h-58, f"Period: {fd}  to  {td}   |   Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}")
        # Summary Cards
        y = h - 110
        c.setFillColor("black"); c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, "SUMMARY")
        y -= 20
        items = [
            ("Total Sales",     self._data["sales"],     "#2ECC71"),
            ("Total Purchases", self._data["purchases"], "#2980B9"),
            ("Total Expenses",  self._data["expenses"],  "#F39C12"),
            ("Net Profit/Loss", self._data["net"],
             "#27AE60" if self._data["net"] >= 0 else "#E74C3C"),
        ]
        box_w = (w - 80) / 4
        for i, (title, val, color) in enumerate(items):
            x = 40 + i * box_w
            # Box
            c.setFillColor(color)
            c.roundRect(x, y-50, box_w-10, 55, 5, fill=1, stroke=0)
            c.setFillColor("white")
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x+8, y-15, title)
            c.setFont("Helvetica-Bold", 13)
            c.drawString(x+8, y-38, f"Rs{val:,.2f}")
        y -= 80
        # Detail tables
        def draw_table(title, color, rows, headers):
            nonlocal y
            if not rows: return
            if y < 120: c.showPage(); y = h - 60
            c.setFillColor(color); c.rect(40, y-2, w-80, 20, fill=1, stroke=0)
            c.setFillColor("white"); c.setFont("Helvetica-Bold", 11)
            c.drawString(44, y+3, title)
            y -= 22
            # Headers
            c.setFillColor("#2C3E50"); c.rect(40, y-2, w-80, 18, fill=1, stroke=0)
            c.setFillColor("white"); c.setFont("Helvetica-Bold", 8)
            col_x = [44, 120, 220, 330, 390, 440, 490]
            for j, hdr in enumerate(headers):
                if j < len(col_x): c.drawString(col_x[j], y+2, hdr)
            y -= 18
            c.setFont("Helvetica", 8)
            for ri, row in enumerate(rows):
                if y < 60: c.showPage(); y = h-60; c.setFont("Helvetica", 8)
                bg = "#F7F9FC" if ri%2==0 else "white"
                c.setFillColor(bg); c.rect(40, y-2, w-80, 16, fill=1, stroke=0)
                c.setFillColor("#2C3E50")
                for j, val in enumerate(row):
                    if j < len(col_x): c.drawString(col_x[j], y+1, str(val)[:18])
                y -= 16
            y -= 10

        # Get detail data
        conn = get_connection()
        sales_rows = conn.execute("""
            SELECT t.date,p.name,pr.name,t.entered_quantity,
                   t.price_per_unit,t.gst,t.total_amount
            FROM transactions t
            JOIN parties p ON p.id=t.party_id
            JOIN products pr ON pr.id=t.product_id
            WHERE t.type='SALE' AND t.date BETWEEN ? AND ?
            ORDER BY t.date""", (fd,td)).fetchall()
        purch_rows = conn.execute("""
            SELECT t.date,p.name,pr.name,t.entered_quantity,
                   t.price_per_unit,t.gst,t.total_amount
            FROM transactions t
            JOIN parties p ON p.id=t.party_id
            JOIN products pr ON pr.id=t.product_id
            WHERE t.type='PURCHASE' AND t.date BETWEEN ? AND ?
            ORDER BY t.date""", (fd,td)).fetchall()
        exp_rows = conn.execute("""
            SELECT e.date,COALESCE(emp.name,'-'),e.category,
                   e.description,e.amount
            FROM expenses e
            LEFT JOIN employees emp ON emp.id=e.employee_id
            WHERE e.date BETWEEN ? AND ?
            ORDER BY e.date""", (fd,td)).fetchall()
        conn.close()

        hdrs = ["Date","Party","Product","Qty","Rate","GST","Total"]
        draw_table("SALES", "#2ECC71",
                   [(r[0],r[1][:15],r[2][:15],f"{r[3]:.1f}",
                     f"{r[4]:.2f}",f"{r[5]:.2f}",f"Rs{r[6]:.2f}") for r in sales_rows], hdrs)
        draw_table("PURCHASES", "#2980B9",
                   [(r[0],r[1][:15],r[2][:15],f"{r[3]:.1f}",
                     f"{r[4]:.2f}",f"{r[5]:.2f}",f"Rs{r[6]:.2f}") for r in purch_rows], hdrs)
        draw_table("EXPENSES", "#F39C12",
                   [(r[0],r[1][:15],r[2][:12],r[3][:15] if r[3] else "-",
                     f"Rs{r[4]:.2f}","","") for r in exp_rows],
                   ["Date","Employee","Category","Description","Amount","",""])
        c.save()
        messagebox.showinfo("PDF", f"PDF saved!\n{path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  PRODUCTION LEDGER
# ═══════════════════════════════════════════════════════════════════════════════
class ProductionLedgerFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._rows_cache = []; self._build()

    def _build(self):
        SectionHeader(self, "Production Ledger").pack(fill="x")
        conn = get_connection()
        self._products = [dict(r) for r in conn.execute("SELECT id,name FROM products ORDER BY name").fetchall()]
        conn.close()
        filt = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=PAD_SMALL); filt.pack(fill="x")
        tk.Label(filt,text="From:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left")
        self.from_var = tk.StringVar(value="2020-01-01")
        make_date_entry(filt, self.from_var).pack(side="left", padx=4)
        tk.Label(filt,text="To:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left", padx=(8,0))
        self.to_var = tk.StringVar(value=today_str())
        make_date_entry(filt, self.to_var).pack(side="left", padx=4)
        tk.Label(filt,text="Product:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left", padx=(8,0))
        self.prod_var = tk.StringVar(value="All")
        ttk.Combobox(filt,textvariable=self.prod_var,values=["All"]+[p["name"] for p in self._products],
                     state="readonly",width=20,font=FONT_NORMAL).pack(side="left", padx=4)
        StyledButton(filt,"📊 Generate",command=self._generate,kind="info").pack(side="left", padx=8)
        if EXCEL_OK: StyledButton(filt,"⬆ Excel",command=self._export,kind="neutral").pack(side="left", padx=(0,6))
        if PDF_OK:   StyledButton(filt,"📄 PDF",  command=self._pdf,   kind="danger" ).pack(side="left")

        self.table = StyledTable(self, [("batch","Batch",70),("date","Date",90),("prod","Product",160),
            ("inp","Input Qty",100),("out","Output Qty",100),("loss","Loss %",80),("bal","Balance",100)])
        # Search bar
        self.search_bar = SearchBar(self, self.table, "🔍  Product ya batch se search karein...")
        self.search_bar.pack(fill="x", padx=PAD, pady=(4,0))
        self.table.pack(fill="both", expand=True, padx=PAD, pady=PAD_SMALL)

    def _generate(self):
        fd=self.from_var.get(); td=self.to_var.get(); pn=self.prod_var.get()
        conn = get_connection()
        batches = conn.execute("SELECT id,date FROM production_batches WHERE date BETWEEN ? AND ? ORDER BY date",(fd,td)).fetchall()
        data = []
        for b in batches:
            inputs  = conn.execute("""SELECT pr.name, pr.base_unit,
                SUM(pi.quantity_base) as t FROM production_inputs pi
                JOIN products pr ON pr.id=pi.product_id
                WHERE pi.batch_id=? GROUP BY pi.product_id""",(b["id"],)).fetchall()
            outputs = conn.execute("""SELECT pr.name, pr.base_unit,
                SUM(po.quantity_base) as t FROM production_outputs po
                JOIN products pr ON pr.id=po.product_id
                WHERE po.batch_id=? GROUP BY po.product_id""",(b["id"],)).fetchall()
            im = {r["name"]:(r["t"], r["base_unit"]) for r in inputs}
            om = {r["name"]:(r["t"], r["base_unit"]) for r in outputs}
            for pname in sorted(set(list(im)+list(om))):
                if pn != "All" and pname != pn: continue
                inp, unit = im.get(pname, (0, ""))
                out, _    = om.get(pname, (0, ""))
                balance   = inp - out
                loss_pct  = (balance / inp * 100) if inp > 0 else 0.0
                # Color hint: if output only (no input) balance is negative
                data.append((b["id"], b["date"], pname,
                             f"{inp:.2f} {unit}", f"{out:.2f} {unit}",
                             f"{loss_pct:.1f}%", f"{balance:.2f} {unit}"))
        conn.close()
        self.table.load(data); self._rows_cache = data

    def _export(self):
        if not EXCEL_OK or not self._rows_cache: return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Production Ledger"
        ws.append(["Batch","Date","Product","Input","Output","Loss%","Balance"])
        for r in self._rows_cache: ws.append(list(r))
        wb.save(path); messagebox.showinfo("Export", f"Saved: {path}")

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        if not self._rows_cache: messagebox.showinfo("PDF","Pehle Generate karein."); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF","*.pdf")])
        if not path: return
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        # Header
        c.setFillColor("#1A2332"); c.rect(0, h-70, w, 70, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 16)
        c.drawString(40, h-38, "Saark Industries — Production Ledger")
        c.setFont("Helvetica", 9)
        c.drawString(40, h-56, f"Period: {self.from_var.get()} to {self.to_var.get()}   |   Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}")
        y = h - 90
        # Table headers
        c.setFillColor("#2C3E50"); c.rect(40, y-2, w-80, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 9)
        col_x = [44, 90, 150, 290, 370, 450, 500]
        for i, hdr in enumerate(["Batch","Date","Product","Input Qty","Output Qty","Loss%","Balance"]):
            c.drawString(col_x[i], y+2, hdr)
        y -= 18
        c.setFont("Helvetica", 8)
        for ri, row in enumerate(self._rows_cache):
            if y < 60: c.showPage(); y = h-60; c.setFont("Helvetica", 8)
            bg = "#F7F9FC" if ri%2==0 else "white"
            c.setFillColor(bg); c.rect(40, y-2, w-80, 16, fill=1, stroke=0)
            c.setFillColor("#2C3E50")
            for i, val in enumerate(row):
                if i < len(col_x): c.drawString(col_x[i], y+1, str(val)[:20])
            y -= 16
        # Total rows
        y -= 5
        c.setFillColor("#1A2332"); c.rect(40, y-2, w-80, 18, fill=1, stroke=0)
        c.setFillColor("white"); c.setFont("Helvetica-Bold", 9)
        c.drawString(44, y+2, f"Total Batches: {len(set(r[0] for r in self._rows_cache))}   |   Total Records: {len(self._rows_cache)}")
        c.save()
        messagebox.showinfo("PDF", f"PDF saved!\n{path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  EXPENSES
# ═══════════════════════════════════════════════════════════════════════════════
CATEGORIES = ["Electricity","Transport","Maintenance","Other","Salary"]

class ExpenseFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG_MAIN)
        self._load_employees(); self._build(); self.load_data()

    def _load_employees(self):
        conn = get_connection()
        self._employees = [dict(r) for r in conn.execute("SELECT id,name FROM employees ORDER BY name").fetchall()]
        conn.close()

    def _build(self):
        SectionHeader(self, "Expense Management").pack(fill="x")
        form = tk.Frame(self, bg=BG_CARD, padx=PAD, pady=PAD,
                        highlightthickness=1, highlightbackground=BORDER)
        form.pack(fill="x", padx=PAD, pady=PAD_SMALL)
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        # Row 0: Date + Category
        tk.Label(form,text="Date *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=0,column=0,sticky="w",pady=6,padx=(0,8))
        self.date_var = tk.StringVar(value=today_str())
        make_date_entry(form, self.date_var).grid(row=0,column=1,sticky="w",padx=4)
        tk.Label(form,text="Category *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=0,column=2,sticky="w",pady=6,padx=(16,8))
        self.cat_var = tk.StringVar(value="Other")
        cc = ttk.Combobox(form,textvariable=self.cat_var,values=CATEGORIES,state="readonly",width=20,font=FONT_NORMAL)
        cc.grid(row=0,column=3,sticky="ew",padx=4); cc.bind("<<ComboboxSelected>>",self._on_cat)

        # Row 1: Employee (full width with + Emp button)
        tk.Label(form,text="Employee",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=1,column=0,sticky="w",pady=6,padx=(0,8))
        emp_frame = tk.Frame(form, bg=BG_CARD)
        emp_frame.grid(row=1,column=1,columnspan=2,sticky="ew",padx=4)
        self.emp_var = tk.StringVar()
        self.emp_cb = ttk.Combobox(emp_frame,textvariable=self.emp_var,values=[e["name"] for e in self._employees],state="disabled",width=28,font=FONT_NORMAL)
        self.emp_cb.pack(side="left", padx=(0,8))
        self.emp_var.trace_add("write",self._on_emp)
        StyledButton(emp_frame,"+ Add Employee",command=self._add_emp,kind="neutral").pack(side="left")

        # Row 1 right: Amount
        tk.Label(form,text="Amount *",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=1,column=3,sticky="w",pady=6,padx=(16,4))

        # Row 2: Amount entry (separate row to avoid overlap)
        self.amt_var = tk.StringVar()
        tk.Entry(form,textvariable=self.amt_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1,width=20).grid(row=2,column=3,sticky="ew",padx=4,ipady=5)

        # Row 2: Description
        tk.Label(form,text="Description",font=FONT_SMALL,bg=BG_CARD,fg=TEXT_MUTED).grid(row=2,column=0,sticky="w",pady=6,padx=(0,8))
        self.desc_var = tk.StringVar()
        tk.Entry(form,textvariable=self.desc_var,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1).grid(row=2,column=1,columnspan=2,sticky="ew",padx=4,ipady=5)

        # Row 3: Buttons
        btn_row = tk.Frame(form, bg=BG_CARD)
        btn_row.grid(row=3,column=0,columnspan=4,pady=10,sticky="w")
        StyledButton(btn_row,"💾 Save",command=self._save).pack(side="left",padx=(0,8))
        StyledButton(btn_row,"🗑 Delete",command=self._delete,kind="danger").pack(side="left",padx=(0,8))
        if EXCEL_OK: StyledButton(btn_row,"⬆ Excel",command=self._export,kind="info").pack(side="left",padx=(0,8))
        if PDF_OK:   StyledButton(btn_row,"📄 PDF",command=self._pdf,kind="neutral").pack(side="left")

        # Filter row - Category + Employee + Month
        filt = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=4); filt.pack(fill="x")

        tk.Label(filt,text="From:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left")
        self.f_from = tk.StringVar(value="2020-01-01")
        make_date_entry(filt, self.f_from).pack(side="left", padx=4)
        tk.Label(filt,text="To:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left", padx=(8,0))
        self.f_to = tk.StringVar(value=today_str())
        make_date_entry(filt, self.f_to).pack(side="left", padx=4)

        # Category filter
        tk.Label(filt,text="Category:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left", padx=(12,4))
        self.f_cat = tk.StringVar(value="All")
        ttk.Combobox(filt,textvariable=self.f_cat,
                     values=["All"]+CATEGORIES,
                     state="readonly",width=14,font=FONT_NORMAL).pack(side="left", padx=4)

        # Employee filter
        tk.Label(filt,text="Employee:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left", padx=(8,4))
        self.f_emp = tk.StringVar(value="All")
        self.f_emp_cb = ttk.Combobox(filt,textvariable=self.f_emp,
                                      values=["All"]+[e["name"] for e in self._employees],
                                      state="readonly",width=16,font=FONT_NORMAL)
        self.f_emp_cb.pack(side="left", padx=4)

        StyledButton(filt,"🔍 Filter",command=self.load_data,kind="info").pack(side="left", padx=8)
        StyledButton(filt,"↺ Reset",command=self._reset_filter,kind="neutral").pack(side="left")

        # Description filter - alag row mein
        filt2 = tk.Frame(self, bg=BG_MAIN, padx=PAD, pady=2); filt2.pack(fill="x")
        tk.Label(filt2,text="Description:",font=FONT_SMALL,bg=BG_MAIN,fg=TEXT_MUTED).pack(side="left")
        self.f_desc = tk.StringVar()
        self.f_desc.trace_add("write", lambda *_: self.load_data())
        desc_entry = tk.Entry(filt2, textvariable=self.f_desc, font=FONT_NORMAL,
                              bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1, width=40)
        desc_entry.pack(side="left", padx=6, ipady=4)
        # Clear button
        tk.Button(filt2, text="✕", font=FONT_SMALL, bg=ENTRY_BG, fg=TEXT_MUTED,
                  relief="flat", bd=0, cursor="hand2",
                  command=lambda: self.f_desc.set("")).pack(side="left")
        tk.Label(filt2, text="← type karo, apne aap filter ho jayega",
                 font=FONT_SMALL, bg=BG_MAIN, fg=TEXT_MUTED).pack(side="left", padx=8)

        self.table = StyledTable(self, [("id","ID",40),("date","Date",90),("cat","Category",110),
            ("emp","Employee",140),("desc","Description",200),("amt","Amount",100)])
        # Search bar
        self.search_bar = SearchBar(self, self.table, "🔍  Category, description ya employee se search karein...")
        self.search_bar.pack(fill="x", padx=PAD, pady=(4,0))
        self.table.pack(fill="both", expand=True, padx=PAD, pady=4)
        self.total_lbl = tk.Label(self, text="Total: ₹0.00", font=FONT_SUBTITLE, bg=BG_MAIN, fg=DANGER)
        self.total_lbl.pack(anchor="e", padx=PAD, pady=4)

    def _on_cat(self, _=None):
        self.emp_cb.config(state="readonly" if self.cat_var.get()=="Salary" else "disabled")
        if self.cat_var.get() != "Salary": self.emp_var.set("")
    def _on_emp(self, *_):
        if self.cat_var.get()=="Salary" and self.emp_var.get():
            self.desc_var.set(f"Salary of {self.emp_var.get()}")
    def _add_emp(self): AddEmployeeDialog(self, on_save=self._refresh_emp)
    def _refresh_emp(self):
        self._load_employees()
        self.emp_cb["values"] = [e["name"] for e in self._employees]
        self.f_emp_cb["values"] = ["All"] + [e["name"] for e in self._employees]

    def _reset_filter(self):
        self.f_from.set("2020-01-01")
        self.f_to.set(today_str())
        self.f_cat.set("All")
        self.f_emp.set("All")
        self.f_desc.set("")
        self.load_data()

    def load_data(self):
        conn = get_connection()
        # Build query with filters
        filters = ["e.date BETWEEN ? AND ?"]
        params  = [self.f_from.get(), self.f_to.get()]

        if self.f_cat.get() != "All":
            filters.append("e.category=?")
            params.append(self.f_cat.get())

        if self.f_emp.get() != "All":
            emp = next((e for e in self._employees if e["name"] == self.f_emp.get()), None)
            if emp:
                filters.append("e.employee_id=?")
                params.append(emp["id"])

        # Description filter
        desc_kw = self.f_desc.get().strip()
        if desc_kw:
            filters.append("LOWER(COALESCE(e.description,'')) LIKE ?")
            params.append(f"%{desc_kw.lower()}%")

        where = " AND ".join(filters)
        rows = conn.execute(f"""
            SELECT e.id, e.date, e.category,
                   COALESCE(emp.name,'') as en,
                   e.description, e.amount
            FROM expenses e
            LEFT JOIN employees emp ON emp.id=e.employee_id
            WHERE {where}
            ORDER BY e.date DESC""", params).fetchall()
        conn.close()
        total = sum(r["amount"] for r in rows)
        self.table.load([(r["id"],r["date"],r["category"],r["en"],
                          r["description"] or "",f"{r['amount']:.2f}") for r in rows])
        # Show filtered total
        cat_txt  = self.f_cat.get()
        emp_txt  = self.f_emp.get()
        desc_txt = self.f_desc.get().strip()
        label    = "Total"
        if cat_txt  != "All": label += f" ({cat_txt})"
        if emp_txt  != "All": label += f" [{emp_txt}]"
        if desc_txt:          label += f" | '{desc_txt}'"
        self.total_lbl.config(text=f"{label}: ₹{total:,.2f}")

    def _save(self):
        date=self.date_var.get(); cat=self.cat_var.get(); desc=self.desc_var.get().strip() or None
        try: amt = float(self.amt_var.get())
        except ValueError: messagebox.showerror("Error","Amount must be number."); return
        emp_id = None
        if cat == "Salary":
            en = self.emp_var.get()
            if not en: messagebox.showerror("Error","Select employee for Salary."); return
            emp = next((e for e in self._employees if e["name"]==en), None)
            if emp: emp_id = emp["id"]
        conn = get_connection()
        conn.execute("INSERT INTO expenses (date,category,employee_id,description,amount) VALUES (?,?,?,?,?)",
                     (date,cat,emp_id,desc,amt))
        conn.commit(); conn.close()
        self.load_data(); self.amt_var.set(""); self.desc_var.set("")

    def _delete(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Delete","Select an expense."); return
        if not messagebox.askyesno("Delete","Delete this expense?",parent=self): return
        conn = get_connection()
        conn.execute("DELETE FROM expenses WHERE id=?", (sel[0],)); conn.commit(); conn.close(); self.load_data()

    def _export(self):
        if not EXCEL_OK: return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")])
        if not path: return
        conn = get_connection()
        rows = conn.execute("SELECT e.date,e.category,COALESCE(emp.name,''),e.description,e.amount FROM expenses e LEFT JOIN employees emp ON emp.id=e.employee_id ORDER BY e.date DESC").fetchall()
        conn.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
        ws.append(["Date","Category","Employee","Description","Amount"])
        for r in rows: ws.append(list(r))
        wb.save(path); messagebox.showinfo("Export",f"Saved: {path}")

    def _pdf(self):
        if not PDF_OK: messagebox.showinfo("PDF","pip install reportlab"); return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF","*.pdf")])
        if not path: return
        conn = get_connection()
        rows = conn.execute("SELECT e.date,e.category,COALESCE(emp.name,'') as emp,e.description,e.amount FROM expenses e LEFT JOIN employees emp ON emp.id=e.employee_id ORDER BY e.date DESC").fetchall()
        conn.close()
        c = rl_canvas.Canvas(path, pagesize=A4); w, h = A4
        c.setFont("Helvetica-Bold",16); c.drawString(50,h-60,"Saark Industries – Expense Report")
        c.setFont("Helvetica",10); c.drawString(50,h-80,f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        y = h-110; col_x=[50,130,220,310,460]
        c.setFont("Helvetica-Bold",10)
        for i,h_ in enumerate(["Date","Category","Employee","Description","Amount"]): c.drawString(col_x[i],y,h_)
        y-=15; c.line(50,y,w-50,y); y-=10; c.setFont("Helvetica",9); total=0.0
        for r in rows:
            if y<60: c.showPage(); y=h-60; c.setFont("Helvetica",9)
            for i,v in enumerate([r[0],r[1],r[2],r[3] or "",f"Rs{r[4]:.2f}"]): c.drawString(col_x[i],y,str(v)[:25])
            total+=r[4]; y-=18
        y-=5; c.line(50,y,w-50,y); y-=15; c.setFont("Helvetica-Bold",10); c.drawString(380,y,f"Total: Rs{total:.2f}")
        c.save(); messagebox.showinfo("PDF",f"Saved: {path}")


class AddEmployeeDialog(ModalDialog):
    def __init__(self, parent, on_save=None):
        super().__init__(parent, "Add Employee", 400, 260)
        self.on_save = on_save; self._build()

    def _build(self):
        b = self.body; b.columnconfigure(1,weight=1)
        self.vars = {}
        for i,(lbl,key) in enumerate([("Name *","name"),("Position","pos"),("Salary","sal")]):
            tk.Label(b,text=lbl,font=FONT_NORMAL,bg=BG_CARD,fg=TEXT_MAIN).grid(row=i,column=0,sticky="w",pady=6,padx=(0,12))
            v = tk.StringVar(); self.vars[key] = v
            tk.Entry(b,textvariable=v,font=FONT_NORMAL,bg=ENTRY_BG,fg=ENTRY_FG,relief="solid",bd=1).grid(row=i,column=1,sticky="ew",ipady=4)
        r = tk.Frame(b,bg=BG_CARD); r.grid(row=3,column=0,columnspan=2,pady=12)
        StyledButton(r,"💾 Save",command=self._save).pack(side="left",padx=(0,8))
        StyledButton(r,"Cancel",command=self.destroy,kind="neutral").pack(side="left")

    def _save(self):
        name=self.vars["name"].get().strip(); pos=self.vars["pos"].get().strip() or None
        try: sal=float(self.vars["sal"].get() or 0)
        except ValueError: sal=0
        if not name: messagebox.showerror("Error","Name required.",parent=self); return
        conn = get_connection()
        conn.execute("INSERT INTO employees (name,position,salary) VALUES (?,?,?)", (name,pos,sal))
        conn.commit(); conn.close()
        if self.on_save: self.on_save()
        self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  USERS MANAGEMENT (sirf admin ke liye)
# ═══════════════════════════════════════════════════════════════════════════════
class UsersFrame(tk.Frame):
    def __init__(self, parent, current_user=None):
        super().__init__(parent, bg=BG_MAIN)
        self._current_user = current_user or {}
        self._build(); self.load_data()

    def _build(self):
        SectionHeader(self, "👥 Users Management").pack(fill="x")

        # Warning banner
        warn = tk.Frame(self, bg="#FEF9E7", padx=PAD, pady=8,
                        highlightthickness=1, highlightbackground=WARNING)
        warn.pack(fill="x", padx=PAD, pady=(0, PAD_SMALL))
        tk.Label(warn, text="⚠  Sirf Admin log yahan access kar sakte hain!  "
                            "Apna Developer account kabhi delete mat karna.",
                 font=FONT_SMALL, bg="#FEF9E7", fg=WARNING_DARK).pack(side="left")

        bar = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); bar.pack(fill="x")
        StyledButton(bar, "+ Add User", command=self._add).pack(side="left")

        self.table = StyledTable(self, [
            ("id",       "ID",        40),
            ("username", "Username", 200),
            ("role",     "Role",     100),
            ("created",  "Created",  160),
        ])
        self.table.pack(fill="both", expand=True, padx=PAD, pady=PAD_SMALL)

        b = tk.Frame(self, bg=BG_MAIN, pady=PAD_SMALL, padx=PAD); b.pack(fill="x")
        StyledButton(b, "✏ Edit / Reset Password", command=self._edit,   kind="warning").pack(side="left", padx=(0,8))
        StyledButton(b, "🗑 Delete User",           command=self._delete, kind="danger").pack(side="left")
        tk.Label(b, text="(Admin account delete mat karna!)",
                 font=FONT_SMALL, bg=BG_MAIN, fg=TEXT_MUTED).pack(side="left", padx=12)

    def load_data(self):
        conn = get_connection()
        rows = conn.execute(
            "SELECT id, username, role, created_at FROM users ORDER BY id"
        ).fetchall()
        conn.close()
        self.table.load([(r["id"], r["username"], r["role"],
                          r["created_at"][:16]) for r in rows])

    def _add(self):    UserDialog(self, on_save=self.load_data)
    def _edit(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Edit", "Pehle ek user select karein."); return
        conn = get_connection()
        row = conn.execute("SELECT * FROM users WHERE id=?", (sel[0],)).fetchone()
        conn.close()
        if row: UserDialog(self, record=dict(row), on_save=self.load_data)

    def _delete(self):
        sel = self.table.get_selected()
        if not sel: messagebox.showinfo("Delete", "Pehle ek user select karein."); return

        # Khud apna account delete nahi kar sakta
        if sel[0] == self._current_user.get("id"):
            messagebox.showerror("Error ❌",
                "Aap apna khud ka account delete nahi kar sakte!\n\n"
                "Dusre admin se delete karwao."); return

        # Last admin delete nahi ho sakta
        conn = get_connection()
        count = conn.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]
        sel_role = conn.execute("SELECT role FROM users WHERE id=?", (sel[0],)).fetchone()
        conn.close()
        if count <= 1 and sel_role and sel_role["role"] == "admin":
            messagebox.showerror("Error ❌",
                "Sirf ek Admin bacha hai!\n\n"
                "Pehle koi aur Admin banao,\n"
                "phir is account ko delete karo."); return

        if not messagebox.askyesno("Delete",
                f"'{sel[1]}' user delete karein?\n\nYe user dobara login nahi kar payega.",
                parent=self): return
        conn = get_connection()
        conn.execute("DELETE FROM users WHERE id=?", (sel[0],))
        conn.commit(); conn.close()
        self.load_data()
        messagebox.showinfo("Done ✅", f"User '{sel[1]}' delete ho gaya!")


class UserDialog(ModalDialog):
    def __init__(self, parent, record=None, on_save=None):
        title = "Add New User" if not record else "Edit User / Reset Password"
        super().__init__(parent, title, 440, 380)
        self.record = record; self.on_save = on_save
        self._build()
        if record: self._populate()

    def _build(self):
        b = self.body; b.columnconfigure(1, weight=1)

        # Info label
        info_txt = ("Naya user banao — sirf yahi log app use kar payenge!" if not self.record
                    else "Password change karo ya role update karo.")
        tk.Label(b, text=info_txt, font=FONT_SMALL, bg=BG_CARD,
                 fg=TEXT_MUTED, wraplength=380).grid(row=0, column=0, columnspan=2,
                                                      sticky="w", pady=(0,12))

        fields = [("Username *", "uname"), ("Password *", "pwd"),
                  ("Confirm Password *", "cpwd")]
        self.vars = {}
        for i, (lbl, key) in enumerate(fields):
            tk.Label(b, text=lbl, font=FONT_NORMAL, bg=BG_CARD,
                     fg=TEXT_MAIN).grid(row=i+1, column=0, sticky="w", pady=8, padx=(0,12))
            v = tk.StringVar(); self.vars[key] = v
            show = "•" if "pwd" in key else ""
            tk.Entry(b, textvariable=v, show=show, font=FONT_NORMAL,
                     bg=ENTRY_BG, fg=ENTRY_FG, relief="solid", bd=1
                     ).grid(row=i+1, column=1, sticky="ew", ipady=4)

        # Role
        tk.Label(b, text="Role", font=FONT_NORMAL, bg=BG_CARD,
                 fg=TEXT_MAIN).grid(row=4, column=0, sticky="w", pady=8, padx=(0,12))
        self.role_var = tk.StringVar(value="user")
        role_frame = tk.Frame(b, bg=BG_CARD); role_frame.grid(row=4, column=1, sticky="w")
        tk.Radiobutton(role_frame, text="user  (sirf data entry)",
                       variable=self.role_var, value="user",
                       bg=BG_CARD, fg=TEXT_MAIN, font=FONT_SMALL).pack(anchor="w")
        tk.Radiobutton(role_frame, text="admin (sab kuch + Users manage)",
                       variable=self.role_var, value="admin",
                       bg=BG_CARD, fg=DANGER, font=FONT_SMALL).pack(anchor="w")

        # Edit note
        if self.record:
            tk.Label(b, text="💡 Password khali chodo agar change nahi karna",
                     font=FONT_SMALL, bg=BG_CARD, fg=ACCENT
                     ).grid(row=5, column=0, columnspan=2, sticky="w", pady=(4,0))

        r = tk.Frame(b, bg=BG_CARD); r.grid(row=6, column=0, columnspan=2, pady=16)
        StyledButton(r, "💾 Save", command=self._save).pack(side="left", padx=(0,8))
        StyledButton(r, "Cancel", command=self.destroy, kind="neutral").pack(side="left")

    def _populate(self):
        self.vars["uname"].set(self.record["username"])
        self.role_var.set(self.record["role"])

    def _save(self):
        uname = self.vars["uname"].get().strip()
        pwd   = self.vars["pwd"].get().strip()
        cpwd  = self.vars["cpwd"].get().strip()
        role  = self.role_var.get()

        if not uname:
            messagebox.showerror("Error", "Username zaroori hai.", parent=self); return

        # New user — password required
        if not self.record and not pwd:
            messagebox.showerror("Error", "Naye user ke liye password zaroori hai.", parent=self); return

        # Password validation
        if pwd:
            if len(pwd) < 6:
                messagebox.showerror("Error", "Password kam se kam 6 characters ka hona chahiye.", parent=self); return
            if pwd != cpwd:
                messagebox.showerror("Error", "Dono passwords same nahi hain!", parent=self); return

        try:
            conn = get_connection()
            if self.record:
                if pwd:
                    conn.execute("UPDATE users SET username=?, role=?, password=? WHERE id=?",
                                 (uname, role, hash_password(pwd), self.record["id"]))
                else:
                    conn.execute("UPDATE users SET username=?, role=? WHERE id=?",
                                 (uname, role, self.record["id"]))
            else:
                # Check duplicate username
                exists = conn.execute("SELECT id FROM users WHERE username=?", (uname,)).fetchone()
                if exists:
                    messagebox.showerror("Error", f"'{uname}' username already exist karta hai!", parent=self)
                    conn.close(); return
                conn.execute("INSERT INTO users (username, password, role) VALUES (?,?,?)",
                             (uname, hash_password(pwd), role))
            conn.commit(); conn.close()
            messagebox.showinfo("Success",
                                f"User '{uname}' {'update' if self.record else 'add'} ho gaya! ✅",
                                parent=self)
            if self.on_save: self.on_save()
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL BACKUP SETTINGS DIALOG
# ═══════════════════════════════════════════════════════════════════════════════
class EmailBackupDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("📧 Email Backup Settings")
        self.resizable(True, True)
        self.configure(bg=BG_CARD)
        self.grab_set()
        w, h = 520, 480
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w = min(w, sw - 40)
        h = min(h, sh - 80)
        px = parent.winfo_rootx() + parent.winfo_width()  // 2 - w // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - h // 2
        px = max(10, min(px, sw - w - 10))
        py = max(10, min(py, sh - h - 10))
        self.geometry(f"{w}x{h}+{px}+{py}")
        self._build()
        self._load()

    def _build(self):
        tk.Frame(self, height=4, bg="#3498DB").pack(fill="x")
        tk.Label(self, text="📧 Email Backup Settings", font=FONT_SUBTITLE,
                 bg=BG_CARD, fg=TEXT_MAIN, padx=PAD, pady=PAD).pack(anchor="w")
        tk.Frame(self, height=1, bg=BORDER).pack(fill="x")

        body = tk.Frame(self, bg=BG_CARD, padx=PAD*2, pady=PAD)
        body.pack(fill="both", expand=True)
        body.columnconfigure(1, weight=1)

        # How it works info
        info = tk.Frame(body, bg="#EBF5FB", padx=10, pady=8)
        info.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0,12))
        tk.Label(info, text="ℹ  Kaise kaam karta hai:",
                 font=FONT_BOLD, bg="#EBF5FB", fg=INFO).pack(anchor="w")
        tk.Label(info,
                 text="App band karte waqt backup automatically\n"
                      "client ki email pe chali jayegi!\n"
                      "Gmail App Password chahiye (neeche guide hai).",
                 font=FONT_SMALL, bg="#EBF5FB", fg=TEXT_MAIN,
                 justify="left").pack(anchor="w")

        # Fields
        fields = [
            ("Backup bhejni hai is email pe:", "backup_email",  False),
            ("Sender Gmail (app ki):",          "sender_email",  False),
            ("Gmail App Password:",             "sender_password", True),
        ]
        self.vars = {}
        for i, (lbl, key, secret) in enumerate(fields):
            tk.Label(body, text=lbl, font=FONT_SMALL, bg=BG_CARD,
                     fg=TEXT_MUTED).grid(row=i+1, column=0, sticky="w",
                                         pady=6, padx=(0,12))
            v = tk.StringVar(); self.vars[key] = v
            tk.Entry(body, textvariable=v, show="•" if secret else "",
                     font=FONT_NORMAL, bg=ENTRY_BG, fg=ENTRY_FG,
                     relief="solid", bd=1, width=32
                     ).grid(row=i+1, column=1, sticky="ew", ipady=4)

        # Auto backup toggle
        self.auto_var = tk.BooleanVar()
        tk.Checkbutton(body, text="✅ App band hone pe automatically email bhejo",
                       variable=self.auto_var, font=FONT_NORMAL,
                       bg=BG_CARD, fg=TEXT_MAIN
                       ).grid(row=4, column=0, columnspan=2, sticky="w", pady=8)

        # Gmail App Password guide
        guide = tk.Frame(body, bg="#FEF9E7", padx=10, pady=8)
        guide.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(4,12))
        tk.Label(guide, text="📖 Gmail App Password kaise banayein:",
                 font=FONT_BOLD, bg="#FEF9E7", fg=WARNING_DARK).pack(anchor="w")
        tk.Label(guide,
                 text="1. myaccount.google.com pe jao\n"
                      "2. Security → 2-Step Verification ON karo\n"
                      "3. App Passwords → 'Mail' select karo\n"
                      "4. Generate karo → 16-digit password aayega\n"
                      "5. Wahi password yahan daalo",
                 font=FONT_SMALL, bg="#FEF9E7", fg=TEXT_MAIN,
                 justify="left").pack(anchor="w")

        # Buttons
        btn_row = tk.Frame(body, bg=BG_CARD)
        btn_row.grid(row=6, column=0, columnspan=2, pady=12)
        StyledButton(btn_row, "💾 Save", command=self._save).pack(side="left", padx=(0,8))
        StyledButton(btn_row, "📧 Test Email", command=self._test, kind="info").pack(side="left", padx=(0,8))
        StyledButton(btn_row, "Cancel", command=self.destroy, kind="neutral").pack(side="left")

    def _load(self):
        self.vars["backup_email"].set(get_setting("backup_email"))
        self.vars["sender_email"].set(get_setting("sender_email"))
        self.vars["sender_password"].set(get_setting("sender_password"))
        self.auto_var.set(get_setting("auto_email_backup") == "1")

    def _save(self):
        set_setting("backup_email",      self.vars["backup_email"].get().strip())
        set_setting("sender_email",      self.vars["sender_email"].get().strip())
        set_setting("sender_password",   self.vars["sender_password"].get().strip())
        set_setting("auto_email_backup", "1" if self.auto_var.get() else "0")
        messagebox.showinfo("Saved ✅",
                            "Settings save ho gayi!\n\n"
                            "'Test Email' dabao yeh check karne ke liye\n"
                            "ki email sahi ja rahi hai ya nahi.",
                            parent=self)

    def _test(self):
        """Test email bhejo abhi."""
        import shutil, threading
        self._save()
        # Create temp backup for test
        import tempfile
        tmp = tempfile.mktemp(suffix=".db")
        try: shutil.copy2(DB_PATH, tmp)
        except: tmp = DB_PATH

        def do_test():
            ok, err = send_email_backup(tmp)
            if ok:
                messagebox.showinfo("Test Successful! ✅",
                                    f"Email send ho gayi!\n"
                                    f"Check karein: {get_setting('backup_email')}",
                                    parent=self)
            else:
                messagebox.showerror("Test Failed ❌", err, parent=self)

        threading.Thread(target=do_test, daemon=True).start()
        messagebox.showinfo("Sending...", "Email bhej raha hoon...\nThodi der mein pata chalega!", parent=self)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION WINDOW
# ═══════════════════════════════════════════════════════════════════════════════
class MainApp(tk.Tk):
    def __init__(self, user):
        super().__init__()
        self.user = user
        self.title("Saark Industries - ERP")
        try: self.state("zoomed")
        except: self.geometry("1280x800")
        try: self.attributes("-zoomed", True)
        except: pass
        self.configure(bg=BG_MAIN)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._frames = {}; self._active_btn = None
        self._build(); self._show("Dashboard")

    def _build(self):
        self.configure(bg=BG_MAIN)
        # Top bar
        top = tk.Frame(self, bg=BG_TOPBAR, height=TOPBAR_H); top.pack(fill="x"); top.pack_propagate(False)
        tk.Label(top, text="⚙  Saark Industries - ERP", font=FONT_SUBTITLE,
                 bg=BG_TOPBAR, fg=TEXT_MAIN, padx=PAD).pack(side="left", pady=8)
        tk.Label(top, text=f"👤 {self.user['username']}", font=FONT_NORMAL,
                 bg=BG_TOPBAR, fg=TEXT_MUTED).pack(side="right", padx=PAD)
        tk.Frame(top, width=1, bg=BORDER).pack(side="right", fill="y", pady=6)
        for text, cmd in [("🔒 Logout", self._logout), ("🔑 Change Password", self._change_pw)]:
            b = tk.Button(top, text=text, command=cmd, font=FONT_SMALL, bg=BG_TOPBAR, fg=TEXT_MUTED,
                          relief="flat", bd=0, cursor="hand2", pady=4, padx=10,
                          activebackground=BG_MAIN, activeforeground=TEXT_MAIN)
            b.pack(side="right")
            b.bind("<Enter>", lambda e, btn=b: btn.config(fg=TEXT_MAIN))
            b.bind("<Leave>", lambda e, btn=b: btn.config(fg=TEXT_MUTED))
        # Backup / Restore buttons
        tk.Frame(top, width=1, bg=BORDER).pack(side="right", fill="y", pady=6)
        restore_btn = tk.Button(top, text="📂 Restore", command=self._restore_backup,
                                font=FONT_SMALL, bg="#E74C3C", fg="white",
                                relief="flat", bd=0, cursor="hand2", pady=4, padx=10)
        restore_btn.pack(side="right", padx=2)
        backup_btn = tk.Button(top, text="💾 Backup", command=self._take_backup,
                               font=FONT_SMALL, bg="#27AE60", fg="white",
                               relief="flat", bd=0, cursor="hand2", pady=4, padx=10)
        backup_btn.pack(side="right", padx=2)
        # Admin only buttons
        if self.user.get("role") == "admin":
            tk.Frame(top, width=1, bg=BORDER).pack(side="right", fill="y", pady=6)
            email_btn = tk.Button(top, text="📧 Email Backup", command=lambda: EmailBackupDialog(self),
                                  font=FONT_SMALL, bg="#3498DB", fg="white",
                                  relief="flat", bd=0, cursor="hand2", pady=4, padx=10)
            email_btn.pack(side="right", padx=2)
            users_btn = tk.Button(top, text="👥 Users", command=lambda: self._show("Users"),
                                  font=FONT_SMALL, bg="#F39C12", fg="white",
                                  relief="flat", bd=0, cursor="hand2", pady=4, padx=10)
            users_btn.pack(side="right", padx=2)
        # Dark/Light mode toggle
        self._theme_btn_text = tk.StringVar(value="🌙 Dark Mode" if _current_theme=="light" else "☀ Light Mode")
        theme_btn = tk.Button(top, textvariable=self._theme_btn_text, command=self._toggle_theme,
                              font=FONT_SMALL, bg="#6C5CE7" if _current_theme=="light" else "#F39C12",
                              fg="white", relief="flat", bd=0, cursor="hand2", pady=4, padx=12)
        theme_btn.pack(side="right", padx=6)
        self._theme_btn = theme_btn
        tk.Frame(top, width=1, bg=BORDER).pack(side="right", fill="y", pady=6)
        tk.Frame(self, height=1, bg=BORDER).pack(fill="x")

        # Main container
        main = tk.Frame(self, bg=BG_MAIN); main.pack(fill="both", expand=True)

        # Sidebar
        sidebar = tk.Frame(main, bg=BG_SIDEBAR, width=SIDEBAR_W); sidebar.pack(fill="y", side="left"); sidebar.pack_propagate(False)
        tk.Frame(sidebar, height=8, bg=ACCENT).pack(fill="x")
        tk.Label(sidebar, text="MENU", font=(FONT_FAMILY,8,"bold"), bg=BG_SIDEBAR,
                 fg=TEXT_SIDEBAR, pady=10).pack(anchor="w", padx=16)

        self._sb_btns = {}
        for label, key in [
            ("📊  Dashboard","Dashboard"),("📦  Products","Products"),("🤝  Parties","Parties"),
            ("🔁  Transactions","Transactions"),("🏭  Production","Production"),
            ("📒  Party Ledger","PartyLedger"),("💰  Profit / Loss","ProfitLoss"),
            ("📋  Production Ledger","ProductionLedger"),("🧾  Expenses","Expenses")]:
            btn = tk.Button(sidebar, text=label, anchor="w", font=FONT_SIDEBAR,
                            bg=BG_SIDEBAR, fg=TEXT_SIDEBAR, relief="flat", bd=0,
                            cursor="hand2", padx=20, pady=10, width=22,
                            activebackground=SIDEBAR_HOVER_BG, activeforeground=TEXT_LIGHT,
                            command=lambda k=key: self._show(k))
            btn.pack(fill="x")
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg=SIDEBAR_HOVER_BG, fg=TEXT_LIGHT) if b != self._active_btn else None)
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg=BG_SIDEBAR, fg=TEXT_SIDEBAR) if b != self._active_btn else None)
            self._sb_btns[key] = btn

        # Content
        self.content = tk.Frame(main, bg=BG_MAIN); self.content.pack(fill="both", expand=True)
        self.content.rowconfigure(0, weight=1); self.content.columnconfigure(0, weight=1)

    def _show(self, key):
        if key not in self._frames:
            frame_map = {
                "Dashboard":         DashboardFrame,
                "Products":          ProductsFrame,
                "Parties":           PartiesFrame,
                "Transactions":      TransactionsFrame,
                "Production":        ProductionFrame,
                "PartyLedger":       PartyLedgerFrame,
                "ProfitLoss":        ProfitLossFrame,
                "ProductionLedger":  ProductionLedgerFrame,
                "Expenses":          ExpenseFrame,
                "Users":             lambda p: UsersFrame(p, self.user),
            }
            cls = frame_map.get(key)
            if cls: self._frames[key] = cls(self.content)
            else:
                f = tk.Frame(self.content, bg=BG_MAIN)
                tk.Label(f, text=f"{key} – Coming Soon", font=FONT_SUBTITLE, bg=BG_MAIN, fg=TEXT_MUTED).pack(pady=40)
                self._frames[key] = f

        for f in self._frames.values(): f.grid_remove()
        self._frames[key].grid(row=0, column=0, sticky="nsew")

        if self._active_btn: self._active_btn.config(bg=BG_SIDEBAR, fg=TEXT_SIDEBAR)
        if key in self._sb_btns:
            self._active_btn = self._sb_btns[key]
            self._active_btn.config(bg=SIDEBAR_ACTIVE_BG, fg=SIDEBAR_ACTIVE_TEXT)

        if key == "Dashboard": self._frames[key].refresh()

    def _change_pw(self): ChangePasswordDialog(self, self.user)

    def _take_backup(self):
        """Manual backup — user chooses where to save."""
        import shutil
        from tkinter.filedialog import asksaveasfilename
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        default  = f"saark_backup_{date_str}.db"
        path = asksaveasfilename(
            defaultextension=".db",
            initialfile=default,
            filetypes=[("Database Backup","*.db"),("All Files","*.*")],
            title="Backup kahan save karein?"
        )
        if not path: return
        try:
            shutil.copy2(DB_PATH, path)
            messagebox.showinfo("Backup Successful! ✅",
                f"Backup save ho gaya:\n{path}\n\n"
                f"Ye file safe jagah rakhein — Pen Drive ya Google Drive mein!")
        except Exception as e:
            messagebox.showerror("Backup Failed", str(e))

    def _restore_backup(self):
        """Restore from backup file."""
        import shutil
        from tkinter.filedialog import askopenfilename
        if not messagebox.askyesno("Restore Backup ⚠",
            "Restore karne se ABHI KA SARA DATA HAT JAYEGA!\n"
            "Aur backup wala data aa jayega.\n\n"
            "Kya aap sure hain?", parent=self):
            return
        path = askopenfilename(
            filetypes=[("Database Backup","*.db"),("All Files","*.*")],
            title="Backup file select karein"
        )
        if not path: return
        try:
            # Auto backup current data before restore
            import shutil as sh
            date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
            auto_bkp = DB_PATH.replace(".db", f"_before_restore_{date_str}.db")
            sh.copy2(DB_PATH, auto_bkp)
            # Restore
            sh.copy2(path, DB_PATH)
            messagebox.showinfo("Restore Successful! ✅",
                f"Data restore ho gaya!\n\n"
                f"Purana data save hai:\n{auto_bkp}\n\n"
                f"App restart ho rahi hai...")
            # Restart app
            user = self.user
            self.destroy()
            apply_theme(_current_theme)
            initialize_database()
            app = MainApp(user)
            app.mainloop()
        except Exception as e:
            messagebox.showerror("Restore Failed", str(e))

    def _auto_backup(self):
        """Auto backup on app close — last 7 days rakho + email bhejo."""
        import shutil, glob, threading
        try:
            bkp_folder = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "backups")
            os.makedirs(bkp_folder, exist_ok=True)
            date_str = datetime.now().strftime("%Y-%m-%d")
            bkp_path = os.path.join(bkp_folder, f"saark_auto_{date_str}.db")
            shutil.copy2(DB_PATH, bkp_path)
            # Purane backups delete karo (7 din se purane)
            all_bkps = sorted(glob.glob(os.path.join(bkp_folder, "saark_auto_*.db")))
            if len(all_bkps) > 7:
                for old in all_bkps[:-7]:
                    try: os.remove(old)
                    except: pass
            # Email backup agar enabled hai
            if get_setting("auto_email_backup") == "1":
                def send_bg():
                    ok, err = send_email_backup(bkp_path)
                    if not ok:
                        print(f"[Email Backup] Failed: {err}")
                threading.Thread(target=send_bg, daemon=True).start()
        except Exception:
            pass

    def _toggle_theme(self):
        new_theme = "dark" if _current_theme == "light" else "light"
        apply_theme(new_theme)
        user = self.user
        self.destroy()
        app = MainApp(user)
        app.mainloop()

    def _logout(self):
        if messagebox.askyesno("Logout","Logout?",parent=self):
            self._auto_backup()
            self.destroy(); run_app()

    def _on_close(self):
        if messagebox.askyesno("Exit","Exit application?",parent=self):
            self._auto_backup()
            self.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════
def run_app():
    # ── Hardware Lock Check ───────────────────────────────────────────────────
    if not check_hardware_lock():
        show_lock_error()
        return
    # ─────────────────────────────────────────────────────────────────────────
    apply_theme(_current_theme)
    initialize_database()

    # ── First Time Setup Check ────────────────────────────────────────────────
    conn = get_connection()
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()

    if user_count == 0:
        # Pehli baar — setup screen dikhao
        sw = FirstSetupWindow()
        sw.mainloop()
        if not sw.setup_done:
            # User ne setup cancel kiya — band karo
            return

    # ── Normal Login ──────────────────────────────────────────────────────────
    lw = LoginWindow()
    lw.mainloop()
    if lw.logged_in_user:
        app = MainApp(lw.logged_in_user)
        app.mainloop()


if __name__ == "__main__":
    run_app()
